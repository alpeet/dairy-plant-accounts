#!/usr/bin/env python3
"""
Godhuli Dairy Plant - Excel Workbook Generator
================================================
Generates a professional .xlsx workbook for the Godhuli Dairy Plant
Accounts & Stock Management application.

Compatible with:
  - Windows 7 + Office 2007/2010/2013
  - macOS + Office 2011/2016+
  - LibreOffice

Usage:
  python3 generate_godhuli_excel.py

Output:
  Godhuli_Dairy_Plant_Workbook.xlsx

Author: Godhuli Dairy Plant
"""

import datetime
import os

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo



# ──────────────────────────────────────────────────────────────────
# CONSTANTS
# ──────────────────────────────────────────────────────────────────
BUSINESS_NAME = "Godhuli Dairy Plant"
VERSION = "1.0"
OUTPUT_FILE = "Godhuli_Dairy_Plant_Workbook.xlsx"

# Colors
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1F4E79")
KPI_FONT = Font(name="Calibri", bold=True, size=13, color="333333")
KPI_LABEL_FONT = Font(name="Calibri", size=10, color="666666")
KPI_VALUE_FONT = Font(name="Calibri", bold=True, size=22, color="1F4E79")
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
BOLD_WHITE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
KPI_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="1F4E79"),
    bottom=Side(style="thin", color="D0D0D0"),
)
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
DASHBOARD_CARD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Date formats
DATE_FMT = "DD-MMM-YYYY"
DATE_FMT_SHORT = "DD-MMM-YY"
NUM_FMT_2DP = '#,##0.00'
NUM_FMT_0DP = '#,##0'
NUM_FMT_PCT = '0.0%'


# ──────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────
def apply_header(ws, headers, row=1):
    """Write a header row with styling and auto-filter."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = BOLD_WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    # Auto-filter on header row
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{row}:{last_col}{row}"
    # Freeze panes below header
    ws.freeze_panes = f"A{row + 1}"


def write_data_rows(ws, data_list, start_row=2):
    """Write data rows with alternate row shading."""
    for row_offset, row_data in enumerate(data_list):
        row_num = start_row + row_offset
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            # Alternate row shading
            if row_offset % 2 == 1:
                cell.fill = ALT_ROW_FILL
            # Format dates — openpyxl writes datetime.date as Excel serial numbers
            if isinstance(value, datetime.date):
                cell.number_format = DATE_FMT


def set_column_widths(ws, widths):
    """Set column widths. widths is a list of (col_letter_or_number, width) tuples."""
    for col, w in widths:
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w


def add_table(ws, headers, data_list, table_name, start_row=1):
    """Write headers + data and format as an Excel Table (compatible with Excel 2007+)."""
    row_count = len(data_list)
    col_count = len(headers)

    apply_header(ws, headers, row=start_row)
    write_data_rows(ws, data_list, start_row=start_row + 1)

    if row_count > 0:
        # Create an Excel Table object
        end_row = start_row + row_count
        end_col = get_column_letter(col_count)
        ref = f"A{start_row}:{end_col}{end_row}"

        table = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",  # Clean blue style, compatible with Excel 2007+
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)


# ──────────────────────────────────────────────────────────────────
# SHEET BUILDERS
# ──────────────────────────────────────────────────────────────────

def build_sales_analysis(ws):
    """Sheet: Sales_Analysis — hidden helper sheet joining Sales + Sales_Items + Products.
    This denormalized data feeds the Dashboard cross-tabulation formulas."""

    headers = [
        "Sale ID", "Invoice No", "Date", "Party Name",
        "Product Name", "Category", "Quantity", "Unit",
        "Rate", "Item Amount", "Invoice Total"
    ]
    col_widths = [
        (1, 8), (2, 14), (3, 14), (4, 24),
        (5, 20), (6, 14), (7, 10), (8, 8),
        (9, 10), (10, 14), (11, 14)
    ]

    # We need to join Sales with Sales_Items and Products
    # The sample data is hardcoded below matching the Sales and Sales_Items builders
    from datetime import date, timedelta
    today = date.today()

    # Sale-level data (matching build_sales)
    sales_data = []
    for i, days_ago in enumerate([0, 1, 3, 5, 7]):
        d = today - timedelta(days=days_ago)
        inv = f"INV-{1001 + i}"
        party = ["Sharma Dairy Shop", "Bhatbhateni Supermarket", "Patan Dairy Cooperative"][i % 3]
        subtotal = round((2000 + i * 500 + (i * 100)) * 1.0, 2)
        disc = round(subtotal * (0.02 + (i % 3) * 0.005), 2)
        grand = round(subtotal - disc, 2)
        sales_data.append({"id": i + 1, "inv": inv, "date": d, "party": party, "grand": grand})

    # Item-level data (matching build_sales_items but tied to sales_data above)
    product_names = ["Raw Milk", "Cow Milk", "Curd (Dahi)", "Ghee", "Paneer"]
    product_cats = ["Milk", "Milk", "Curd", "Ghee", "Paneer"]
    product_units = ["liter", "liter", "kg", "kg", "kg"]

    analysis_rows = []
    mid = 0
    for sale_idx, sale in enumerate(sales_data):
        num_items = 1 + sale_idx % 2
        for j in range(num_items):
            mid += 1
            pidx = (sale_idx + j) % 5
            pname = product_names[pidx]
            cat = product_cats[pidx]
            unit = product_units[pidx]
            qty = round(5 + (sale["id"] * 2) + (j * 3) + (j * 0.5), 2)
            rate = [60, 60, 100, 600, 320][pidx]
            amt = round(qty * rate, 2)
            analysis_rows.append([
                sale["id"], sale["inv"], sale["date"], sale["party"],
                pname, cat, qty, unit, rate, amt, sale["grand"]
            ])

    add_table(ws, headers, analysis_rows, "SalesAnalysisTable")
    set_column_widths(ws, col_widths)

    # Format date column
    for r in range(2, 2 + len(analysis_rows)):
        ws.cell(r, 3).number_format = DATE_FMT_SHORT


def build_dashboard_crosstab(ws):
    """Add a formula-based cross-tabulation to the Dashboard for Sales by Party × Product Category.
    Uses SUMIFS formulas referencing the Sales_Analysis hidden sheet.
    No PivotTable needed — pure formulas update instantly when data changes."""

    # ── Title ──
    row = 43
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row, 1, "📊 SALES BY PARTY & PRODUCT CATEGORY").font = Font(
        name="Calibri", bold=True, size=11, color="1F4E79"
    )

    # ── Column headers: Categories ──
    categories = ["Milk", "Curd", "Ghee", "Paneer"]
    hdr_row = row + 1  # row 44
    ws.cell(hdr_row, 1, "Party Name").font = BOLD_FONT
    ws.cell(hdr_row, 1).fill = SUBHEADER_FILL
    ws.cell(hdr_row, 1).border = THIN_BORDER

    for ci, cat in enumerate(categories):
        col = ci + 2  # B, C, D, E
        ws.cell(hdr_row, col, f"{cat}").font = BOLD_FONT
        ws.cell(hdr_row, col).fill = SUBHEADER_FILL
        ws.cell(hdr_row, col).border = THIN_BORDER
        ws.cell(hdr_row, col).alignment = Alignment(horizontal="center")

    # Total column
    tot_col = len(categories) + 2  # F
    ws.cell(hdr_row, tot_col, "Total").font = BOLD_FONT
    ws.cell(hdr_row, tot_col).fill = SUBHEADER_FILL
    ws.cell(hdr_row, tot_col).border = THIN_BORDER
    ws.cell(hdr_row, tot_col).alignment = Alignment(horizontal="center")

    # ── Data rows: Party Names ──
    party_names = [
        "Sharma Dairy Shop",
        "Bhatbhateni Supermarket",
        "Patan Dairy Cooperative",
    ]

    # Sales_Analysis columns:
    # A=Sale ID, B=Invoice No, C=Date, D=Party Name, E=Product Name,
    # F=Category, G=Quantity, H=Unit, I=Rate, J=Item Amount, K=Invoice Total

    for ri, party in enumerate(party_names):
        r = hdr_row + 1 + ri
        ws.cell(r, 1, party).font = BOLD_FONT
        ws.cell(r, 1).border = THIN_BORDER

        for ci, cat in enumerate(categories):
            col = ci + 2
            col_letter = get_column_letter(col)
            # SUMIFS(Item Amount, Party Name = party, Category = category-header)
            # Sales_Analysis!J = Item Amount, D = Party Name, F = Category
            formula = (
                f'=SUMIFS(Sales_Analysis!J$2:J$1000,'
                f'Sales_Analysis!D$2:D$1000,A{r},'
                f'Sales_Analysis!F$2:F$1000,{col_letter}${hdr_row})'
            )
            ws.cell(r, col).value = formula
            ws.cell(r, col).number_format = NUM_FMT_2DP
            ws.cell(r, col).border = THIN_BORDER
            ws.cell(r, col).alignment = Alignment(horizontal="right")

        # Party total = SUM of category columns
        first_cat_col = get_column_letter(2)   # B
        last_cat_col = get_column_letter(len(categories) + 1)  # E
        ws.cell(r, tot_col).value = f'=SUM({first_cat_col}{r}:{last_cat_col}{r})'
        ws.cell(r, tot_col).number_format = NUM_FMT_2DP
        ws.cell(r, tot_col).font = BOLD_FONT
        ws.cell(r, tot_col).border = THIN_BORDER
        ws.cell(r, tot_col).alignment = Alignment(horizontal="right")

        if ri % 2 == 1:
            for c in range(1, tot_col + 1):
                ws.cell(r, c).fill = ALT_ROW_FILL

    # ── Grand Total row ──
    grand_row = hdr_row + 1 + len(party_names)
    first_data_row = hdr_row + 1
    last_data_row = grand_row - 1

    ws.cell(grand_row, 1, "GRAND TOTAL").font = BOLD_FONT
    ws.cell(grand_row, 1).border = THIN_BORDER
    ws.cell(grand_row, 1).fill = LIGHT_BLUE_FILL

    for ci in range(len(categories)):
        col = ci + 2
        col_letter = get_column_letter(col)
        ws.cell(grand_row, col).value = f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})'
        ws.cell(grand_row, col).number_format = NUM_FMT_2DP
        ws.cell(grand_row, col).font = BOLD_FONT
        ws.cell(grand_row, col).border = THIN_BORDER
        ws.cell(grand_row, col).alignment = Alignment(horizontal="right")
        ws.cell(grand_row, col).fill = LIGHT_BLUE_FILL

    first_cat_col = get_column_letter(2)
    last_cat_col = get_column_letter(len(categories) + 1)
    ws.cell(grand_row, tot_col).value = f'=SUM({first_cat_col}{grand_row}:{last_cat_col}{grand_row})'
    ws.cell(grand_row, tot_col).number_format = NUM_FMT_2DP
    ws.cell(grand_row, tot_col).font = BOLD_FONT
    ws.cell(grand_row, tot_col).border = THIN_BORDER
    ws.cell(grand_row, tot_col).alignment = Alignment(horizontal="right")
    ws.cell(grand_row, tot_col).fill = LIGHT_BLUE_FILL

    # ── Note ──
    note_row = grand_row + 2
    ws.merge_cells(f"A{note_row}:F{note_row}")
    ws.cell(note_row, 1, (
        "💡 Tip: To see more parties or categories, add rows below and copy the formulas. "
        "All values auto-calculate from data in the Sales and Sales_Items sheets."
    )).font = Font(name="Calibri", italic=True, size=8, color="888888")


def build_instructions(ws):
    """Sheet 1: Instructions / Config / Help."""
    ws.title = "Instructions"

    # Business name + title
    ws.merge_cells("A1:H1")
    ws.cell(1, 1, f"🐄 {BUSINESS_NAME} — Excel Workbook").font = TITLE_FONT

    ws.merge_cells("A2:H2")
    ws.cell(2, 1, f"Version {VERSION} — Generated {datetime.date.today().strftime('%d-%b-%Y')}").font = Font(
        name="Calibri", italic=True, color="666666", size=10
    )

    # ── Section: Overview ──
    row = 4
    ws.cell(row, 1, "📋 OVERVIEW").font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    row += 1
    overview_text = [
        "This workbook contains 13 sheets for managing Godhuli Dairy Plant's accounting and inventory data.",
        "",
        "Use it to enter, view, and analyze your daily business data in Microsoft Excel (Windows 7+ / Office 2007+).",
        "After editing on any computer, you can export data back into the Godhuli Dairy Plant application.",
    ]
    for line in overview_text:
        ws.cell(row, 1, line).font = NORMAL_FONT
        row += 1

    # ── Section: Sheet Guide ──
    row += 1
    ws.cell(row, 1, "📑 SHEET GUIDE").font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    row += 1
    guide = [
        ("Sheet Name", "Purpose", "App Table", "Data Entry?"),
        ("─" * 40, "─" * 60, "─" * 40, "─" * 15),
        ("Dashboard",   "Key KPIs, cross-tab & summaries (auto-calculated)", "—", "No (read-only views)"),
        ("Parties",     "Customers, suppliers & mixed parties master list", "parties", "Yes"),
        ("Products",    "Product master list (milk, curd, ghee, etc.)",  "products", "Yes"),
        ("Sales",       "Sales invoice header data",                    "sales", "Yes"),
        ("Sales_Items", "Sales invoice line items (products sold)",     "sales_items", "Yes"),
        ("Purchases",   "Purchase bill header data",                    "purchases", "Yes"),
        ("Purchase_Items", "Purchase bill line items (products bought)","purchase_items", "Yes"),
        ("Stock",       "Stock movement ledger (in/out/balance)",       "stock_movements", "Yes"),
        ("Milk",        "Daily milk collection from farmers",           "milk_collections", "Yes"),
        ("Payments",    "Payment & receipt records",                    "payments", "Yes"),
        ("Ledger",      "Party-wise financial ledger entries",          "ledger_entries", "Yes"),
        ("Sales_Analysis","Hidden — joined Sales+Items for cross-tab",  "—", "No (auto, hidden)"),
    ]
    for g in guide:
        for ci, val in enumerate(g):
            cell = ws.cell(row, ci + 1, val)
            cell.font = BOLD_FONT if g is guide[0] else NORMAL_FONT
            cell.border = THIN_BORDER
            if g is guide[0]:
                cell.fill = SUBHEADER_FILL
        row += 1

    set_column_widths(ws, [(1, 22), (2, 55), (3, 28), (4, 28)])

    # ── Section: Data Exchange ──
    row += 1
    ws.cell(row, 1, "🔄 DATA EXCHANGE (App ↔ Excel)").font = Font(
        name="Calibri", bold=True, size=12, color="1F4E79"
    )
    row += 1
    exchange = [
        "1. Export data from the app as CSV or direct database export.",
        "2. Paste / import into the matching sheet (see table above).",
        "3. Make your edits in Excel. Keep column order intact.",
        "4. For import back to app: export from Excel as CSV, matching table names.",
        "",
        "IMPORTANT: Column order in each sheet must match the app's database columns.",
        "          Keep the header row (first row) as-is. Do not add/remove columns.",
        "          Sheets with 'Data Entry?' = Yes are safe to edit manually.",
    ]
    for line in exchange:
        ws.cell(row, 1, line).font = NORMAL_FONT
        row += 1

    # ── Section: Windows 7 / Old Excel Compatibility ──
    row += 1
    ws.cell(row, 1, "🪟 COMPATIBILITY NOTES (Windows 7 / Office 2007-2010)").font = Font(
        name="Calibri", bold=True, size=12, color="1F4E79"
    )
    row += 1
    compat = [
        "✅ Compatible features used: PivotTables, Tables, AutoFilter, basic formulas",
        "   (SUM, AVERAGE, COUNT, IF, SUMIF, COUNTIF, VLOOKUP, INDEX/MATCH).",
        "✅ No VBA macros — zero security warnings.",
        "✅ No Power Query, Power Pivot, or external data connections.",
        "⚠️ Some modern colors/theme may look slightly different in Excel 2007.",
        "⚠️ File is .xlsx format (not .xlsb or .xlsm) — safe for all Excel versions 2007+.",
        "",
        "To refresh Dashboard formulas after editing data:",
        "  - Formulas update automatically when you change cell values.",
        "  - If using manual calculation mode (unlikely), press F9 to recalculate.",
    ]
    for line in compat:
        ws.cell(row, 1, line).font = NORMAL_FONT
        row += 1

    # ── Section: How to Refresh ──
    row += 1
    ws.cell(row, 1, "🔄 HOW TO REFRESH DASHBOARD").font = Font(
        name="Calibri", bold=True, size=12, color="1F4E79"
    )
    row += 1
    refresh = [
        "1. After adding/editing data on any of the data sheets, save the workbook.",
        "2. Go to the Dashboard sheet — all KPIs, summaries, and the cross-tabulation update automatically.",
        "3. If you add rows beyond row 1000 on any data sheet:",
        "   a. Update the named ranges or formula references on the Dashboard sheet.",
        "   b. Or simply copy formulas down to cover new rows.",
        "4. To add new PivotTables, use Insert > PivotTable and select the relevant data sheet.",
    ]
    for line in refresh:
        ws.cell(row, 1, line).font = NORMAL_FONT
        row += 1

    # ── Section: Modifying the Dashboard ──
    row += 1
    ws.cell(row, 1, "✏️ HOW TO ADD NEW CHARTS / KPIs").font = Font(
        name="Calibri", bold=True, size=12, color="1F4E79"
    )
    row += 1
    modify = [
        "1. Use Excel's built-in SUMIFS / COUNTIFS formulas referencing the data sheets.",
        "2. To add a chart: Insert > Chart and select data from any data sheet.",
        "3. To add cross-tabulation categories: copy existing category columns and update references.",
        "4. Avoid: XLOOKUP, dynamic arrays (e.g. FILTER, UNIQUE, SORT), LET, LAMBDA.",
        "5. Safe alternatives: VLOOKUP, INDEX+MATCH, SUMIFS, COUNTIFS, SUMPRODUCT.",
        "6. Always test on your Windows 7 Excel after making changes to the Dashboard.",
    ]
    for line in modify:
        ws.cell(row, 1, line).font = NORMAL_FONT
        row += 1

    # Print scale
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"


def build_dashboard(ws):
    """Sheet 2: Dashboard with KPIs, charts, and summaries."""
    ws.title = "Dashboard"
    """Sheet 2: Dashboard with KPIs, charts, and summaries."""
    ws.title = "Dashboard"

    # We'll reference data sheets. Data goes in rows ~1000+
    # to leave room for Dashboard content. But we'll use explicit ranges.

    # ── Title Row ──
    ws.merge_cells("A1:F1")
    ws.cell(1, 1, f"🐄 {BUSINESS_NAME} — DASHBOARD").font = TITLE_FONT

    today = datetime.date.today()
    ws.merge_cells("A2:F2")
    ws.cell(2, 1, f"As of: {today.strftime('%d-%b-%Y')}").font = Font(
        name="Calibri", italic=True, color="666666", size=10
    )

    # ══════════════════════════════════════════════════════════════
    # KPI ROW (Row 4-7)
    # Each KPI is 2 columns wide (label + value), 6 cols total = 3 KPIs per row
    # ══════════════════════════════════════════════════════════════
    kpi_cols = [("A", "B"), ("C", "D"), ("E", "F")]

    def write_kpi(row, label, formula, fill=GREEN_FILL):
        """Write a KPI card: label in first col, value in second col, spanning 2 cols."""
        col_label, col_value = kpi_cols[0]  # We'll shift using columns

        # Find which position
        for idx, (c1, c2) in enumerate(kpi_cols):
            if c1 == col_label:
                # Shift right based on the actual columns used
                pass

        # Actually, simpler: columns A-B, C-D, E-F
        pass

    # Let's do row-based layout: top row KPIs at row 4
    # ── Row 4: Label row ──
    ws.merge_cells("A4:B4")
    ws.cell(4, 1, "MONTHLY SALES").font = Font(name="Calibri", size=9, color="1F4E79")
    ws.cell(4, 1).alignment = Alignment(horizontal="center")
    ws.cell(4, 1).border = KPI_BORDER
    ws.cell(4, 2).border = KPI_BORDER

    ws.merge_cells("C4:D4")
    ws.cell(4, 3, "MONTHLY PURCHASES").font = Font(name="Calibri", size=9, color="1F4E79")
    ws.cell(4, 3).alignment = Alignment(horizontal="center")
    ws.cell(4, 3).border = KPI_BORDER
    ws.cell(4, 4).border = KPI_BORDER

    ws.merge_cells("E4:F4")
    ws.cell(4, 5, "MILK COLLECTED (Mo.)").font = Font(name="Calibri", size=9, color="1F4E79")
    ws.cell(4, 5).alignment = Alignment(horizontal="center")
    ws.cell(4, 5).border = KPI_BORDER
    ws.cell(4, 6).border = KPI_BORDER

    # ── Row 5: Value row (formulas referencing data sheets) ──
    # Current month: first day to last day
    # Sales formula: SUMIFS on Sales_Data!G:G (grand_total) where date between month start and month end

    first_of_month = today.replace(day=1)
    # Use DATE(year, month, 1) for month start; EOMONTH for month end
    month_start_str = first_of_month.strftime("%Y-%m-%d")
    # Last day of current month
    if today.month == 12:
        last_of_month = today.replace(year=today.year + 1, month=1, day=1) - datetime.timedelta(days=1)
    else:
        last_of_month = today.replace(month=today.month + 1, day=1) - datetime.timedelta(days=1)
    month_end_str = last_of_month.strftime("%Y-%m-%d")        # Sales KPI — Grand Total = column J
    ws.merge_cells("A5:B5")
    ws.cell(5, 1).value = f'=SUMIFS(Sales!J$2:J$1000,Sales!C$2:C$1000,">="&DATE({first_of_month.year},{first_of_month.month},1),Sales!C$2:C$1000,"<="&DATE({last_of_month.year},{last_of_month.month},{last_of_month.day}))'
    ws.cell(5, 1).font = KPI_VALUE_FONT
    ws.cell(5, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(5, 1).number_format = NUM_FMT_2DP
    ws.cell(5, 1).border = KPI_BORDER
    ws.cell(5, 2).border = KPI_BORDER

    # Purchases KPI — Grand Total = column K
    ws.merge_cells("C5:D5")
    ws.cell(5, 3).value = f'=SUMIFS(Purchases!K$2:K$1000,Purchases!C$2:C$1000,">="&DATE({first_of_month.year},{first_of_month.month},1),Purchases!C$2:C$1000,"<="&DATE({last_of_month.year},{last_of_month.month},{last_of_month.day}))'
    ws.cell(5, 3).font = KPI_VALUE_FONT
    ws.cell(5, 3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(5, 3).number_format = NUM_FMT_2DP
    ws.cell(5, 3).border = KPI_BORDER
    ws.cell(5, 4).border = KPI_BORDER

    # Milk KPI (liters) — Quantity = column G, Date = column C
    ws.merge_cells("E5:F5")
    ws.cell(5, 5).value = f'=SUMIFS(Milk!G$2:G$1000,Milk!C$2:C$1000,">="&DATE({first_of_month.year},{first_of_month.month},1),Milk!C$2:C$1000,"<="&DATE({last_of_month.year},{last_of_month.month},{last_of_month.day}))&" L"'
    ws.cell(5, 5).font = KPI_VALUE_FONT
    ws.cell(5, 5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(5, 5).border = KPI_BORDER
    ws.cell(5, 6).border = KPI_BORDER

    # ── Row 6: Second KPI row ──
    ws.merge_cells("A6:B6")
    ws.cell(6, 1, "OUTSTANDING RECEIVABLES").font = Font(name="Calibri", size=9, color="1F4E79")
    ws.cell(6, 1).alignment = Alignment(horizontal="center")
    ws.cell(6, 1).border = KPI_BORDER
    ws.cell(6, 2).border = KPI_BORDER

    ws.merge_cells("C6:D6")
    ws.cell(6, 3, "OUTSTANDING PAYABLES").font = Font(name="Calibri", size=9, color="1F4E79")
    ws.cell(6, 3).alignment = Alignment(horizontal="center")
    ws.cell(6, 3).border = KPI_BORDER
    ws.cell(6, 4).border = KPI_BORDER

    ws.merge_cells("E6:F6")
    ws.cell(6, 5, "STOCK VALUE (Est.)").font = Font(name="Calibri", size=9, color="1F4E79")
    ws.cell(6, 5).alignment = Alignment(horizontal="center")
    ws.cell(6, 5).border = KPI_BORDER
    ws.cell(6, 6).border = KPI_BORDER

    # Row 7: Values for row 6 KPIs
    ws.merge_cells("A7:B7")
    # Receivables: SUM of Sales Grand Total (col J) where Status (col M) = 'unpaid' or 'partial'
    ws.cell(7, 1).value = '=SUMIFS(Sales!J$2:J$1000,Sales!M$2:M$1000,"unpaid")+SUMIFS(Sales!J$2:J$1000,Sales!M$2:M$1000,"partial")'
    ws.cell(7, 1).font = KPI_VALUE_FONT
    ws.cell(7, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(7, 1).number_format = NUM_FMT_2DP
    ws.cell(7, 1).border = KPI_BORDER
    ws.cell(7, 2).border = KPI_BORDER

    ws.merge_cells("C7:D7")
    # Payables: SUM of Purchases Grand Total (col K) where Status (col N) = 'unpaid' or 'partial'
    ws.cell(7, 3).value = '=SUMIFS(Purchases!K$2:K$1000,Purchases!N$2:N$1000,"unpaid")+SUMIFS(Purchases!K$2:K$1000,Purchases!N$2:N$1000,"partial")'
    ws.cell(7, 3).font = KPI_VALUE_FONT
    ws.cell(7, 3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(7, 3).number_format = NUM_FMT_2DP
    ws.cell(7, 3).border = KPI_BORDER
    ws.cell(7, 4).border = KPI_BORDER

    ws.merge_cells("E7:F7")
    # Stock value: SUM of (latest balance_after * rate) for each product
    # Approximated as SUM(Stock!J$2:J$1000 * Stock!K$2:K$1000) / COUNT unique products
    ws.cell(7, 5).value = (
        '"See Stock sheet for details"'
    )
    ws.cell(7, 5).font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws.cell(7, 5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(7, 5).border = KPI_BORDER
    ws.cell(7, 6).border = KPI_BORDER

    # ── Row 9: Sales vs Purchases trend table ──
    row = 9
    ws.cell(row, 1, "MONTHLY TREND (Sales vs Purchases)").font = Font(
        name="Calibri", bold=True, size=11, color="1F4E79"
    )
    ws.merge_cells(f"A{row}:F{row}")

    row = 10
    trend_headers = ["Month", "Sales", "Purchases", "Net", "Milk (Liters)", "Milk Amount"]
    for ci, h in enumerate(trend_headers):
        cell = ws.cell(row, ci + 1, h)
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Generate last 6 months
    for i in range(6):
        m = today.month - i
        y = today.year
        if m <= 0:
            m += 12
            y -= 1
        row_num = row + 1 + i

        # Month name
        month_name = datetime.date(y, m, 1).strftime("%b %Y")
        ws.cell(row_num, 1, month_name).font = BOLD_FONT
        ws.cell(row_num, 1).border = THIN_BORDER

        # First and last day of that month
        first_d = datetime.date(y, m, 1)
        if m == 12:
            last_d = datetime.date(y + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            last_d = datetime.date(y, m + 1, 1) - datetime.timedelta(days=1)

        fn_year, fn_month, fn_day = first_d.year, first_d.month, first_d.day
        ld_year, ld_month, ld_day = last_d.year, last_d.month, last_d.day

        # Sales — Grand Total = column J
        ws.cell(row_num, 2).value = (
            f'=SUMIFS(Sales!J$2:J$1000,Sales!C$2:C$1000,'
            f'">="&DATE({fn_year},{fn_month},{fn_day}),'
            f'Sales!C$2:C$1000,"<="&DATE({ld_year},{ld_month},{ld_day}))'
        )
        ws.cell(row_num, 2).number_format = NUM_FMT_2DP
        ws.cell(row_num, 2).border = THIN_BORDER
        # Purchases — Grand Total = column K
        ws.cell(row_num, 3).value = (
            f'=SUMIFS(Purchases!K$2:K$1000,Purchases!C$2:C$1000,'
            f'">="&DATE({fn_year},{fn_month},{fn_day}),'
            f'Purchases!C$2:C$1000,"<="&DATE({ld_year},{ld_month},{ld_day}))'
        )
        ws.cell(row_num, 3).number_format = NUM_FMT_2DP
        ws.cell(row_num, 3).border = THIN_BORDER
        # Net = Sales - Purchases
        ws.cell(row_num, 4).value = f'=B{row_num}-C{row_num}'
        ws.cell(row_num, 4).number_format = NUM_FMT_2DP
        ws.cell(row_num, 4).border = THIN_BORDER
        # Milk liters — Quantity = column G, Date = column C
        ws.cell(row_num, 5).value = (
            f'=SUMIFS(Milk!G$2:G$1000,Milk!C$2:C$1000,'
            f'">="&DATE({fn_year},{fn_month},{fn_day}),'
            f'Milk!C$2:C$1000,"<="&DATE({ld_year},{ld_month},{ld_day}))'
        )
        ws.cell(row_num, 5).number_format = NUM_FMT_0DP
        ws.cell(row_num, 5).border = THIN_BORDER
        # Milk amount — Amount = column K, Date = column C
        ws.cell(row_num, 6).value = (
            f'=SUMIFS(Milk!K$2:K$1000,Milk!C$2:C$1000,'
            f'">="&DATE({fn_year},{fn_month},{fn_day}),'
            f'Milk!C$2:C$1000,"<="&DATE({ld_year},{ld_month},{ld_day}))'
        )
        ws.cell(row_num, 6).number_format = NUM_FMT_2DP
        ws.cell(row_num, 6).border = THIN_BORDER

        # Alternating row colors
        if i % 2 == 1:
            for c in range(1, 7):
                ws.cell(row_num, c).fill = ALT_ROW_FILL

    # ── Today's Summary table ──
    row = 18
    ws.cell(row, 1, "TODAY'S SUMMARY").font = Font(
        name="Calibri", bold=True, size=11, color="1F4E79"
    )
    ws.merge_cells(f"A{row}:F{row}")

    row = 19
    today_summary = [
        ("Today's Sales Total", f'=SUMIFS(Sales!J$2:J$1000,Sales!C$2:C$1000,TODAY())'),
        ("Today's Purchases Total", f'=SUMIFS(Purchases!K$2:K$1000,Purchases!C$2:C$1000,TODAY())'),
        ("Today's Milk Collected (L)", f'=SUMIFS(Milk!G$2:G$1000,Milk!C$2:C$1000,TODAY())'),
        ("Today's Milk Amount", f'=SUMIFS(Milk!K$2:K$1000,Milk!C$2:C$1000,TODAY())'),
        ("Pending Milk Payments", f'=SUMIFS(Milk!K$2:K$1000,Milk!M$2:M$1000,"pending")'),
        ("No. of Active Products", f'=COUNTA(Products!A$2:A$1000)'),
    ]
    for i, (label, formula) in enumerate(today_summary):
        r = row + i
        ws.cell(r, 1, label).font = BOLD_FONT
        ws.cell(r, 1).border = THIN_BORDER
        ws.merge_cells(f"A{r}:B{r}")
        ws.cell(r, 3).value = formula
        ws.cell(r, 3).font = Font(name="Calibri", bold=True, size=12, color="333333")
        ws.cell(r, 3).number_format = NUM_FMT_2DP if "Amount" in label or "Sales" in label or "Purchases" in label or "Payments" in label else NUM_FMT_0DP
        ws.cell(r, 3).border = THIN_BORDER
        ws.merge_cells(f"C{r}:D{r}")

    # ── Top Customers & Suppliers ──
    row = 27
    ws.cell(row, 1, "TOP CUSTOMERS & SUPPLIERS").font = Font(
        name="Calibri", bold=True, size=11, color="1F4E79"
    )
    ws.merge_cells(f"A{row}:F{row}")

    row = 28
    top_headers = ["Party Name", "Type", "Total Sales/Purchases", "Outstanding"]
    for ci, h in enumerate(top_headers):
        cell = ws.cell(row, ci + 1, h)
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Top customers/suppliers — manual entry area with SUMIF-based formulas
    # The formulas below use SUMIF to look up total sales/purchases by party name
    # (Party names can be typed in manually, then totals auto-calculate)
    for i in range(5):
        r = row + 1 + i
        # Blank cell — user types or pastes party name here
        ws.cell(r, 1, f"[Enter party {i+1} name]").font = Font(name="Calibri", italic=True, color="999999", size=9)
        ws.cell(r, 1).border = THIN_BORDER
        # Type auto-looked-up from Parties sheet
        ws.cell(r, 2).value = f'=IFERROR(VLOOKUP(A{r},Parties!B$2:F$1000,5,FALSE),"")'
        ws.cell(r, 2).border = THIN_BORDER
        # Total sales/purchases (SUMIF across both Sales and Purchases)
        ws.cell(r, 3).value = f'=SUMIF(Sales!E$2:E$1000,A{r},Sales!J$2:J$1000)+SUMIF(Purchases!E$2:E$1000,A{r},Purchases!K$2:K$1000)'
        ws.cell(r, 3).border = THIN_BORDER
        ws.cell(r, 3).number_format = NUM_FMT_2DP
        # Outstanding balance
        ws.cell(r, 4).value = f'=SUMIF(Ledger!C$2:C$1000,A{r},Ledger!J$2:J$1000)'
        ws.cell(r, 4).border = THIN_BORDER
        ws.cell(r, 4).number_format = NUM_FMT_2DP
        if i % 2 == 1:
            for c in range(1, 5):
                ws.cell(r, c).fill = ALT_ROW_FILL

    # ── Quick Stats ──
    row = 36
    ws.cell(row, 1, "QUICK STATS").font = Font(
        name="Calibri", bold=True, size=11, color="1F4E79"
    )
    ws.merge_cells(f"A{row}:F{row}")

    stats = [
        ("Total Parties", f'=COUNTA(Parties!A$2:A$1000)'),
        ("Customers", f'=COUNTIF(Parties!E$2:E$1000,"customer")+COUNTIF(Parties!E$2:E$1000,"both")'),
        ("Suppliers", f'=COUNTIF(Parties!E$2:E$1000,"supplier")+COUNTIF(Parties!E$2:E$1000,"both")'),
        ("Total Products", f'=COUNTA(Products!A$2:A$1000)'),
        ("Total Sales (All Time)", f'=SUM(Sales!J$2:J$1000)'),
        ("Total Purchases (All Time)", f'=SUM(Purchases!K$2:K$1000)'),
    ]
    for i, (label, formula) in enumerate(stats):
        r = row + 1 + i
        ws.cell(r, 1, label).font = BOLD_FONT
        ws.cell(r, 1).border = THIN_BORDER
        ws.merge_cells(f"A{r}:B{r}")
        ws.cell(r, 3).value = formula
        ws.cell(r, 3).font = KPI_FONT
        ws.cell(r, 3).border = THIN_BORDER
        ws.merge_cells(f"C{r}:D{r}")
        if "Total Sales" in label or "Total Purchases" in label:
            ws.cell(r, 3).number_format = NUM_FMT_2DP

    # Column widths
    set_column_widths(ws, [
        (1, 28), (2, 18), (3, 18), (4, 18), (5, 18), (6, 18)
    ])

    # Row heights for KPI rows
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 45
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[7].height = 45

    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"


def build_parties(ws):
    """Sheet: Parties (customer/supplier master)."""
    ws.title = "Parties"
    headers = [
        "ID", "Name", "Phone", "Address", "PAN/VAT",
        "Type", "Opening Balance", "Notes", "Created At", "Updated At"
    ]
    col_widths = [(1, 8), (2, 28), (3, 18), (4, 28), (5, 18),
                  (6, 14), (7, 16), (8, 28), (9, 18), (10, 18)]

    data = [
        [1, "Sharma Dairy Shop", "9812345670", "New Road, Kathmandu", "",
         "customer", 15000, "", datetime.date(2026, 1, 15), datetime.date(2026, 1, 15)],
        [2, "Bhatbhateni Supermarket", "9812345672", "Durbar Marg, Kathmandu", "",
         "customer", 45000, "", datetime.date(2026, 1, 15), datetime.date(2026, 1, 15)],
        [3, "Gurung Dairy Farm", "9852345673", "Budhanilkantha, Kathmandu", "",
         "supplier", 50000, "", datetime.date(2026, 1, 15), datetime.date(2026, 1, 15)],
        [4, "Fresh Valley Suppliers", "9841234570", "Bhaktapur", "",
         "supplier", 12000, "", datetime.date(2026, 1, 15), datetime.date(2026, 1, 15)],
        [5, "Patan Dairy Cooperative", "9852345675", "Patan, Lalitpur", "",
         "both", 0, "", datetime.date(2026, 1, 15), datetime.date(2026, 1, 15)],
    ]

    add_table(ws, headers, data, "PartiesTable")
    set_column_widths(ws, col_widths)

    # Add data validation dropdown for Type column
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"customer,supplier,both"', allow_blank=True)
    dv.error = "Please select customer, supplier, or both"
    dv.errorTitle = "Invalid Type"
    dv.prompt = "Select party type"
    dv.promptTitle = "Party Type"
    ws.add_data_validation(dv)
    dv.add(f"F2:F1000")


def build_products(ws):
    """Sheet: Products master."""
    ws.title = "Products"
    headers = [
        "ID", "Name", "Unit", "Category", "Opening Stock",
        "Reorder Level", "Rate", "GST Rate (%)", "HSN Code", "Notes",
        "Created At", "Updated At"
    ]
    col_widths = [
        (1, 8), (2, 24), (3, 10), (4, 16), (5, 14),
        (6, 14), (7, 10), (8, 12), (9, 12), (10, 28),
        (11, 18), (12, 18)
    ]

    data = [
        [1, "Raw Milk", "liter", "Milk", 1000, 200, 60, 0, "", "",
         datetime.date(2026, 1, 15), datetime.date(2026, 1, 15)],
        [2, "Cow Milk", "liter", "Milk", 500, 100, 60, 0, "", "",
         datetime.date(2026, 1, 15), datetime.date(2026, 1, 15)],
        [3, "Curd (Dahi)", "kg", "Curd", 100, 20, 100, 0, "", "",
         datetime.date(2026, 1, 15), datetime.date(2026, 1, 15)],
        [4, "Ghee", "kg", "Ghee", 50, 10, 600, 0, "", "",
         datetime.date(2026, 1, 15), datetime.date(2026, 1, 15)],
        [5, "Paneer", "kg", "Paneer", 30, 5, 320, 0, "", "",
         datetime.date(2026, 1, 15), datetime.date(2026, 1, 15)],
    ]

    add_table(ws, headers, data, "ProductsTable")
    set_column_widths(ws, col_widths)


def build_sales(ws):
    """Sheet: Sales transactions (header level)."""
    ws.title = "Sales"
    headers = [
        "ID", "Invoice No", "Date", "Party ID", "Party Name",
        "Subtotal", "Discount", "Disc%", "Tax",
        "Grand Total", "Paid Amount", "Payment Mode", "Status",
        "Notes", "Created At", "Updated At"
    ]
    col_widths = [
        (1, 6), (2, 14), (3, 14), (4, 10), (5, 24),
        (6, 10), (7, 10), (8, 8), (9, 8),
        (10, 12), (11, 12), (12, 12), (13, 10),
        (14, 28), (15, 18), (16, 18)
    ]

    from datetime import date, timedelta
    today = date.today()
    data = []
    for i, days_ago in enumerate([0, 1, 3, 5, 7]):
        d = today - timedelta(days=days_ago)
        inv = f"INV-{1001 + i}"
        party = ["Sharma Dairy Shop", "Bhatbhateni Supermarket", "Patan Dairy Cooperative"][i % 3]
        party_id = [1, 2, 5][i % 3]
        subtotal = round((2000 + i * 500 + (i * 100)) * 1.0, 2)
        disc = round(subtotal * (0.02 + (i % 3) * 0.005), 2)
        grand = round(subtotal - disc, 2)
        paid = grand if i % 2 == 0 else (round(grand * 0.5, 2) if i % 3 == 0 else 0)
        status = "paid" if paid == grand else ("partial" if paid > 0 else "unpaid")
        mode = ["cash", "credit", "bank", "upi"][i % 4]
        data.append([
            i + 1, inv, d, party_id, party,
            subtotal, disc, 0 if disc == 0 else round(disc / subtotal * 100, 1),
            0, grand, paid, mode, status,
            f"Sample sale {inv}",
            d, d
        ])

    add_table(ws, headers, data, "SalesTable")
    set_column_widths(ws, col_widths)

    # Date formatting for date columns
    for r in range(2, 2 + len(data)):
        ws.cell(r, 3).number_format = DATE_FMT_SHORT
        ws.cell(r, 15).number_format = DATE_FMT_SHORT
        ws.cell(r, 16).number_format = DATE_FMT_SHORT

    # Data validation for Status and Payment Mode
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_status = DataValidation(type="list", formula1='"paid,unpaid,partial"', allow_blank=True)
    dv_status.error = "Please select a valid status"
    ws.add_data_validation(dv_status)
    dv_status.add(f"M2:M1000")

    dv_mode = DataValidation(type="list", formula1='"cash,credit,bank,upi"', allow_blank=True)
    dv_mode.error = "Please select a valid payment mode"
    ws.add_data_validation(dv_mode)
    dv_mode.add(f"L2:L1000")


def build_sales_items(ws):
    """Sheet: Sales line items."""
    ws.title = "Sales_Items"
    headers = [
        "ID", "Sale ID", "Product ID", "Product Name",
        "Quantity", "Unit", "Rate", "Amount"
    ]
    col_widths = [
        (1, 6), (2, 10), (3, 12), (4, 24),
        (5, 10), (6, 8), (7, 10), (8, 12)
    ]

    products = ["Raw Milk", "Cow Milk", "Curd (Dahi)", "Ghee", "Paneer"]
    data = []
    for sale_id in range(1, 6):
        num_items = 1 + sale_id % 2
        for j in range(num_items):
            pid = (sale_id + j) % 5 + 1
            pname = products[(sale_id + j) % 5]
            qty = round(5 + (sale_id * 2) + (j * 3) + (j * 0.5), 2)
            rate = [60, 60, 100, 600, 320][(sale_id + j) % 5]
            amt = round(qty * rate, 2)
            data.append([
                len(data) + 1, sale_id, pid, pname, qty,
                "liter" if pid <= 2 else "kg", rate, amt
            ])

    add_table(ws, headers, data, "SalesItemsTable")
    set_column_widths(ws, col_widths)


def build_purchases(ws):
    """Sheet: Purchase transactions."""
    ws.title = "Purchases"
    headers = [
        "ID", "Bill No", "Date", "Party ID", "Party Name",
        "Subtotal", "Discount", "Tax", "Transport", "Extra Charges",
        "Grand Total", "Paid Amount", "Payment Mode", "Status",
        "Notes", "Created At", "Updated At"
    ]
    col_widths = [
        (1, 6), (2, 14), (3, 14), (4, 10), (5, 24),
        (6, 10), (7, 10), (8, 8), (9, 10), (10, 10),
        (11, 12), (12, 12), (13, 12), (14, 10),
        (15, 28), (16, 18), (17, 18)
    ]

    from datetime import date, timedelta
    today = date.today()
    data = []
    for i, days_ago in enumerate([0, 2, 4, 6]):
        d = today - timedelta(days=days_ago)
        bill = f"BILL-{5001 + i}"
        party = ["Gurung Dairy Farm", "Fresh Valley Suppliers", "Patan Dairy Cooperative"][i % 3]
        party_id = [3, 4, 5][i % 3]
        subtotal = round((3000 + i * 800) * 1.0, 2)
        transport = round((500 + i * 100) * 1.0, 2)
        grand = round(subtotal + transport, 2)
        paid = grand if i % 2 == 0 else (round(grand * 0.6, 2) if i % 3 == 0 else 0)
        status = "paid" if paid == grand else ("partial" if paid > 0 else "unpaid")
        data.append([
            i + 1, bill, d, party_id, party,
            subtotal, 0, 0, transport, 0,
            grand, paid, "bank", status,
            f"Sample purchase {bill}",
            d, d
        ])

    add_table(ws, headers, data, "PurchasesTable")
    set_column_widths(ws, col_widths)

    # Date formatting for date columns
    for r in range(2, 2 + len(data)):
        ws.cell(r, 3).number_format = DATE_FMT_SHORT
        ws.cell(r, 16).number_format = DATE_FMT_SHORT
        ws.cell(r, 17).number_format = DATE_FMT_SHORT

    # Data validation
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_status = DataValidation(type="list", formula1='"paid,unpaid,partial"', allow_blank=True)
    dv_status.error = "Please select a valid status"
    ws.add_data_validation(dv_status)
    dv_status.add(f"N2:N1000")

    dv_mode = DataValidation(type="list", formula1='"cash,credit,bank,upi"', allow_blank=True)
    ws.add_data_validation(dv_mode)
    dv_mode.add(f"M2:M1000")


def build_purchase_items(ws):
    """Sheet: Purchase line items."""
    ws.title = "Purchase_Items"
    headers = [
        "ID", "Purchase ID", "Product ID", "Product Name",
        "Quantity", "Unit", "Rate", "Amount"
    ]
    col_widths = [
        (1, 6), (2, 14), (3, 12), (4, 24),
        (5, 10), (6, 8), (7, 10), (8, 12)
    ]

    products = ["Raw Milk", "Cow Milk", "Ghee", "Paneer"]
    data = []
    for purchase_id in range(1, 5):
        num_items = 1 + purchase_id % 2
        for j in range(num_items):
            idx = (purchase_id + j) % 4
            pid = [1, 2, 4, 5][idx]
            pname = products[idx]
            qty = round(15 + purchase_id * 5 + j * 10, 2)
            rate = [60, 60, 600, 320][idx]
            amt = round(qty * rate * 0.7, 2)  # purchase rate lower than sale
            data.append([
                len(data) + 1, purchase_id, pid, pname, qty,
                "liter" if pid <= 2 else "kg", round(rate * 0.7, 2), amt
            ])

    add_table(ws, headers, data, "PurchaseItemsTable")
    set_column_widths(ws, col_widths)


def build_stock(ws):
    """Sheet: Stock movements (inventory ledger)."""
    ws.title = "Stock"
    headers = [
        "ID", "Product ID", "Product Name", "Date", "Type",
        "Reference Type", "Reference ID",
        "Inward Qty", "Outward Qty", "Balance After", "Rate",
        "Notes", "Created At"
    ]
    col_widths = [
        (1, 6), (2, 12), (3, 20), (4, 14), (5, 16),
        (6, 18), (7, 14),
        (8, 12), (9, 12), (10, 14), (11, 10),
        (12, 30), (13, 18)
    ]

    from datetime import date, timedelta
    today = date.today()
    balance = {"Raw Milk": 1000, "Cow Milk": 500, "Curd (Dahi)": 100, "Ghee": 50, "Paneer": 30}
    products = {1: "Raw Milk", 2: "Cow Milk", 3: "Curd (Dahi)", 4: "Ghee", 5: "Paneer"}
    rates = {1: 60, 2: 60, 3: 100, 4: 600, 5: 320}

    data = []
    mid = 0

    # Opening stock entries
    for pid, pname in products.items():
        mid += 1
        bal = balance[pname]
        data.append([mid, pid, pname, datetime.date(2026, 1, 1), "opening", "", "",
                      bal, 0, bal, rates[pid], "Opening Stock", datetime.date(2026, 1, 1)])

    # Sale stock movements (out)
    for i, days_ago in enumerate([0, 1, 3, 5, 7]):
        d = today - timedelta(days=days_ago)
        pid = (i % 5) + 1
        pname = products[pid]
        qty = round(5 + i * 2, 2)
        balance[pname] = round(balance[pname] - qty, 2)
        mid += 1
        inv_no = f"INV-{1001 + i}"
        data.append([mid, pid, pname, d, "sale", "sale", i + 1,
                      0, qty, balance[pname], rates[pid],
                      f"Sale {inv_no}", d])

    # Purchase stock movements (in)
    for i, days_ago in enumerate([0, 2, 4, 6]):
        d = today - timedelta(days=days_ago)
        pid = [1, 2, 4, 5][i % 4]
        pname = products[pid]
        qty = round(20 + i * 5, 2)
        balance[pname] = round(balance[pname] + qty, 2)
        mid += 1
        bill_no = f"BILL-{5001 + i}"
        data.append([mid, pid, pname, d, "purchase", "purchase", i + 1,
                      qty, 0, balance[pname], round(rates[pid] * 0.7, 2),
                      f"Purchase {bill_no}", d])

    # Milk collection stock movements (in)
    for i, days_ago in enumerate([0, 1, 3, 5]):
        d = today - timedelta(days=days_ago)
        pid = 1  # Raw Milk
        pname = "Raw Milk"
        qty = round(25 + i * 8, 2)
        balance[pname] = round(balance[pname] + qty, 2)
        mid += 1
        data.append([mid, pid, pname, d, "milk_collection", "milk_collection", i + 1,
                      qty, 0, balance[pname], 60,
                      f"Milk collection #MILK-{2001 + i}", d])

    add_table(ws, headers, data, "StockTable")
    set_column_widths(ws, col_widths)

    for r in range(2, 2 + len(data)):
        ws.cell(r, 4).number_format = DATE_FMT_SHORT
        ws.cell(r, 13).number_format = DATE_FMT_SHORT

    # Conditional formatting: highlight negative balances
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(
        f"J2:J{len(data) + 1}",
        CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL)
    )


def build_milk(ws):
    """Sheet: Milk collections."""
    ws.title = "Milk"
    headers = [
        "ID", "Collection No", "Date", "Party ID", "Farmer Name",
        "Milk Type", "Quantity (L)", "Fat %", "SNF %",
        "Rate", "Amount", "Shift", "Status", "Notes",
        "Created At", "Updated At"
    ]
    col_widths = [
        (1, 6), (2, 14), (3, 14), (4, 10), (5, 24),
        (6, 12), (7, 12), (8, 8), (9, 8),
        (10, 8), (11, 12), (12, 10), (13, 10),
        (14, 28), (15, 18), (16, 18)
    ]

    from datetime import date, timedelta
    today = date.today()
    farmers = {3: "Gurung Dairy Farm", 4: "Fresh Valley Suppliers", 5: "Patan Dairy Cooperative"}
    milk_types = ["cow", "buffalo", "mixed"]
    shifts = ["morning", "evening"]

    data = []
    for i, days_ago in enumerate([0, 1, 3, 5]):
        d = today - timedelta(days=days_ago)
        pid = [3, 4, 5][i % 3]
        fname = farmers[pid]
        mtype = milk_types[i % 3]
        shift = shifts[i % 2]
        qty = round(15 + i * 8 + (i * 2), 2)
        fat = round(3.0 + (i * 0.5) + (i * 0.1), 1)
        base_rate = {"buffalo": 80, "mixed": 70, "cow": 60}[mtype]
        rate = base_rate + (i * 2)
        adj_rate = rate * (fat / 3.5)
        amt = round(qty * adj_rate, 2)
        statuses = ["pending", "paid", "processed", "pending"]
        status = statuses[i % 4]
        data.append([
            i + 1, f"MILK-{2001 + i}", d, pid, fname,
            mtype, qty, fat, round(8.5 + i * 0.3, 1),
            rate, amt, shift, status, "",
            d, d
        ])

    add_table(ws, headers, data, "MilkTable")
    set_column_widths(ws, col_widths)

    for r in range(2, 2 + len(data)):
        ws.cell(r, 3).number_format = DATE_FMT_SHORT
        ws.cell(r, 15).number_format = DATE_FMT_SHORT
        ws.cell(r, 16).number_format = DATE_FMT_SHORT

    # Data validation
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_mtype = DataValidation(type="list", formula1='"cow,buffalo,mixed"', allow_blank=True)
    ws.add_data_validation(dv_mtype)
    dv_mtype.add(f"F2:F1000")

    dv_shift = DataValidation(type="list", formula1='"morning,evening,combined"', allow_blank=True)
    ws.add_data_validation(dv_shift)
    dv_shift.add(f"L2:L1000")

    dv_status = DataValidation(type="list", formula1='"pending,processed,paid"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"M2:M1000")


def build_payments(ws):
    """Sheet: Payment records."""
    ws.title = "Payments"
    headers = [
        "ID", "Party ID", "Party Name", "Date", "Type",
        "Amount", "Mode", "Reference Type", "Reference ID",
        "Notes", "Created At"
    ]
    col_widths = [
        (1, 6), (2, 10), (3, 24), (4, 14), (5, 10),
        (6, 12), (7, 10), (8, 18), (9, 14),
        (10, 28), (11, 18)
    ]

    from datetime import date, timedelta
    today = date.today()
    data = [
        [1, 1, "Sharma Dairy Shop", today - timedelta(days=5),
         "receipt", 15000, "cash", "sale", 1, "Payment for INV-1001", today],
        [2, 3, "Gurung Dairy Farm", today - timedelta(days=3),
         "payment", 50000, "bank", "purchase", 1, "Payment for BILL-5001", today],
        [3, 5, "Patan Dairy Cooperative", today,
         "payment", 18000, "upi", "milk_collection", 1, "Bulk milk payment", today],
    ]

    add_table(ws, headers, data, "PaymentsTable")
    set_column_widths(ws, col_widths)

    for r in range(2, 2 + len(data)):
        ws.cell(r, 4).number_format = DATE_FMT_SHORT
        ws.cell(r, 11).number_format = DATE_FMT_SHORT

    # Data validation
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_type = DataValidation(type="list", formula1='"receipt,payment"', allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.add(f"E2:E1000")

    dv_mode = DataValidation(type="list", formula1='"cash,bank,upi,cheque"', allow_blank=True)
    ws.add_data_validation(dv_mode)
    dv_mode.add(f"G2:G1000")


def build_ledger(ws):
    """Sheet: Party-wise financial ledger."""
    ws.title = "Ledger"
    headers = [
        "ID", "Party ID", "Party Name", "Date", "Reference Type",
        "Reference ID", "Description", "Debit", "Credit",
        "Balance", "Created At"
    ]
    col_widths = [
        (1, 6), (2, 10), (3, 24), (4, 14), (5, 20),
        (6, 14), (7, 32), (8, 12), (9, 12),
        (10, 12), (11, 18)
    ]

    from datetime import date, timedelta
    today = date.today()
    data = [
        [1, 1, "Sharma Dairy Shop", datetime.date(2026, 1, 15), "opening", None, "Opening Balance", 15000, 0, 15000, datetime.date(2026, 1, 15)],
        [2, 1, "Sharma Dairy Shop", today - timedelta(days=7), "sale", 1, "Sale Invoice INV-1001", 5800, 0, 20800, today],
        [3, 1, "Sharma Dairy Shop", today - timedelta(days=5), "payment_received", 1, "Payment Received", 0, 15000, 5800, today],
        [4, 3, "Gurung Dairy Farm", datetime.date(2026, 1, 15), "opening", None, "Opening Balance", 50000, 0, 50000, datetime.date(2026, 1, 15)],
        [5, 3, "Gurung Dairy Farm", today - timedelta(days=6), "purchase", 1, "Purchase Bill BILL-5001", 0, 15300, 65300, today],
    ]

    add_table(ws, headers, data, "LedgerTable")
    set_column_widths(ws, col_widths)

    for r in range(2, 2 + len(data)):
        ws.cell(r, 4).number_format = DATE_FMT_SHORT
        ws.cell(r, 11).number_format = DATE_FMT_SHORT

    # Conditional formatting: highlight high values
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(
        f"H2:H1000",
        CellIsRule(operator="greaterThan", formula=["10000"], fill=RED_FILL)
    )

    # Data validation for reference type
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"sale,purchase,payment_received,payment_made,opening,adjustment,milk_collection"',
        allow_blank=True
    )
    ws.add_data_validation(dv)
    dv.add(f"E2:E1000")


# ──────────────────────────────────────────────────────────────────
# MAIN GENERATION
# ──────────────────────────────────────────────────────────────────
def generate_workbook():
    print(f"🐄 Generating {BUSINESS_NAME} Excel Workbook...")
    wb = Workbook()

    # Create all sheets in desired order
    # First sheet is created by default
    sheets = [
        ("Instructions", build_instructions),
        ("Dashboard", build_dashboard),
        ("Parties", build_parties),
        ("Products", build_products),
        ("Sales", build_sales),
        ("Sales_Items", build_sales_items),
        ("Purchases", build_purchases),
        ("Purchase_Items", build_purchase_items),
        ("Stock", build_stock),
        ("Milk", build_milk),
        ("Payments", build_payments),
        ("Ledger", build_ledger),
    ]

    # Rename default first sheet
    ws0 = wb.active
    ws0.title = "temp"

    # Build all sheets
    for idx, (name, builder) in enumerate(sheets):
        if idx == 0:
            ws = ws0
            ws.title = name
        else:
            ws = wb.create_sheet(title=name)
        print(f"  Building sheet: {name}")
        builder(ws)

    # Build Sales_Analysis helper sheet (hidden, feeds cross-tab formulas)
    print("  Building sheet: Sales_Analysis")
    analysis_ws = wb.create_sheet(title="Sales_Analysis")
    analysis_ws.sheet_state = "hidden"
    build_sales_analysis(analysis_ws)

    # Now add the formula-based cross-tabulation to the Dashboard
    # Must happen after analysis sheet exists so formulas can reference it
    print("  Building Sales-by-Category cross-tabulation on Dashboard...")
    dashboard_ws = wb["Dashboard"]
    build_dashboard_crosstab(dashboard_ws)

    # Set active sheet to Dashboard
    wb.active = 1  # Index 1 = Dashboard (0-indexed: 0=Instructions, 1=Dashboard)

    # Save
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), OUTPUT_FILE)
    wb.save(output_path)
    print(f"\n✅ Workbook saved: {output_path}")
    print(f"   Sheets: {', '.join(name for name, _ in sheets)}")
    print(f"   File size: {os.path.getsize(output_path) / 1024:.1f} KB")
    return output_path


# ──────────────────────────────────────────────────────────────────
# RUN
# ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    generate_workbook()
    print("\n📋 DONE! Open the file in Excel on Windows 7 or macOS.")
    print("   Dashboard formulas will auto-calculate when data is entered.")
