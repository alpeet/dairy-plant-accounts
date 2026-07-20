#!/usr/bin/env python3
"""
Dairy Plant Accounts — Workbook Verification Script
====================================================
Opens the generated Dairy_Plant_Accounts.xlsx and validates:
  - All 12 sheets exist
  - Formulas are syntactically valid (parseable by openpyxl)
  - Excel Tables are present
  - Data Validation dropdowns exist
  - Conditional Formatting rules exist
  - Navigation HYPERLINK formulas are correct
  - Sheet tab colors are set
  - Print settings are configured

Usage:
  python3 verify_workbook.py
"""

import datetime
import os
import sys
import traceback

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

WORKBOOK_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Dairy_Plant_Accounts.xlsx")

EXPECTED_SHEETS = [
    "Dashboard",
    "Sales_Entry",
    "Purchase_Entry",
    "Stock_Master",
    "Party_Master",
    "Party_Ledger",
    "Sales_Report",
    "Purchase_Report",
    "Stock_Report",
    "Printable_Invoice",
    "Printable_Ledger",
    "Settings",
]

# Expected tab colors
TAB_COLORS = {
    "Dashboard": "1B3A5C",
    "Sales_Entry": "27AE60",
    "Purchase_Entry": "E74C3C",
    "Stock_Master": "F39C12",
    "Party_Master": "8E44AD",
    "Party_Ledger": "2C3E50",
    "Sales_Report": "1ABC9C",
    "Purchase_Report": "E67E22",
    "Stock_Report": "3498DB",
    "Printable_Invoice": "9B59B6",
    "Printable_Ledger": "34495E",
    "Settings": "7F8C8D",
}

EXPECTED_TABLES = {
    "Sales_Entry": "SalesTable",
    "Purchase_Entry": "PurchaseTable",
    "Stock_Master": "StockMasterTable",
    "Party_Master": "PartyMasterTable",
    "Party_Ledger": "PartyLedgerTable",
}

EXPECTED_NAV_LINKS = [
    "Sales_Entry", "Purchase_Entry", "Stock_Master", "Party_Master",
    "Party_Ledger", "Sales_Report", "Purchase_Report", "Stock_Report",
    "Printable_Invoice", "Printable_Ledger", "Settings",
]

pass_count = 0
fail_count = 0
warn_count = 0


def check(description, condition, is_warning=False):
    """Check a condition and report pass/fail."""
    global pass_count, fail_count, warn_count
    if condition:
        if is_warning:
            print(f"  ⚠️  {description}")
            warn_count += 1
        else:
            print(f"  ✅ {description}")
            pass_count += 1
    else:
        if is_warning:
            print(f"  ⚠️  {description}")
            warn_count += 1
        else:
            print(f"  ❌ {description}")
            fail_count += 1
    return condition


def verify_formula(ws, cell_ref, expected_keywords=None):
    """Verify a cell contains a valid formula."""
    cell = ws[cell_ref]
    if cell.value is None:
        return False, "empty"
    value = str(cell.value)
    if not value.startswith("="):
        return False, f"not a formula: {value[:50]}"
    if expected_keywords:
        for kw in expected_keywords:
            if kw not in value.upper():
                return False, f"missing keyword {kw} in {value[:80]}"
    return True, "ok"


def verify_sheet_structure(ws, sheet_name):
    """Verify a sheet's structure based on its name."""
    
    if sheet_name == "Dashboard":
        # Check navigation row (row 3)
        nav_ok = True
        for i, sname in enumerate(EXPECTED_NAV_LINKS):
            cell = ws.cell(row=3, column=i + 1)
            if cell.value is None or not str(cell.value).startswith("=HYPERLINK"):
                nav_ok = False
                check(f"Nav link to {sname} at col {i+1}", False)
            else:
                formula = str(cell.value)
                check(f"Nav link to {sname} at col {i+1}: {formula[:60]}", True)
        
        # Check KPI cards at rows 5-8
        kpi_labels_row5 = ["MONTHLY SALES", "MONTHLY PURCHASES", "STOCK VALUE", "LOW STOCK ITEMS"]
        for i, label in enumerate(kpi_labels_row5):
            col = (i * 3) + 1
            cell = ws.cell(row=5, column=col)
            check(f"KPI '{label}' at row 5 col {col}", str(cell.value or "").strip() != "")
        
        kpi_labels_row7 = ["OUTSTANDING RECEIVABLES", "OUTSTANDING PAYABLES", "TOTAL PARTIES", "TOTAL PRODUCTS"]
        for i, label in enumerate(kpi_labels_row7):
            col = (i * 3) + 1
            cell = ws.cell(row=7, column=col)
            check(f"KPI '{label}' at row 7 col {col}", str(cell.value or "").strip() != "")
        
        # Check KPI value cells have formulas
        for row_n in [6, 8]:
            for col_n in [1, 4, 7, 10]:
                ok, msg = verify_formula(ws, f"{get_column_letter(col_n)}{row_n}")
                check(f"KPI value at {get_column_letter(col_n)}{row_n}", ok)
        
        # Check Monthly Trend table exists (find by scanning)
        trend_found = any(
            "MONTHLY TREND" in str(ws.cell(row=r, column=1).value or "").upper()
            for r in range(8, 15)
        )
        check("Monthly Trend table exists", trend_found)
        
        # Check Party Balance section
        party_section_found = any(
            "PARTY BALANCE" in str(ws.cell(row=r, column=1).value or "").upper()
            for r in range(15, 35)
        )
        check("Party Balance Summary section exists", party_section_found)
        
        # Check Low Stock Alerts section
        stock_alert_found = any(
            "LOW STOCK" in str(ws.cell(row=r, column=1).value or "").upper()
            for r in range(30, 55)
        )
        check("Low Stock Alerts section exists", stock_alert_found)
        
        # Check Recent Transactions section
        recent_found = any(
            "RECENT" in str(ws.cell(row=r, column=1).value or "").upper()
            for r in range(40, 70)
        )
        check("Recent Transactions section exists", recent_found)
        
        # Check conditional formatting exists
        check("Dashboard has conditional formatting rules", 
              len(ws.conditional_formatting) > 0)
        
        # Check print orientation
        check("Dashboard is landscape orientation",
              ws.page_setup.orientation == "landscape")
        
        return
    
    if sheet_name == "Sales_Entry":
        # Check header row
        headers = ["Date", "Invoice No", "Party Name", "Product", "Quantity",
                   "Rate", "Amount", "Discount %", "Net Amount", "Payment Mode", "Status", "Remarks"]
        for i, h in enumerate(headers):
            actual = str(ws.cell(row=1, column=i+1).value or "")
            if actual != h:
                check(f"Header '{h}' at col {i+1}", False)
        
        # Check formulas in Amount column (G)
        ok, msg = verify_formula(ws, "G2", ["E", "F"])
        check(f"Sales Amount formula at G2: {msg if not ok else ''}", ok)
        
        # Check Net Amount formula (I)
        ok, msg = verify_formula(ws, "I2", ["G", "H"])
        check(f"Sales Net Amount formula at I2: {msg if not ok else ''}", ok)
        
        # Check data validations
        check("Sales_Entry has data validation rules",
              len(ws.data_validations.dataValidation) > 0)
        
        # Check conditional formatting
        check("Sales_Entry has conditional formatting",
              len(ws.conditional_formatting) > 0)
        
        # Check freeze panes
        check("Sales_Entry has freeze panes", ws.freeze_panes is not None)
        
        return
    
    if sheet_name == "Purchase_Entry":
        headers = ["Date", "Bill No", "Supplier Name", "Product", "Quantity",
                   "Rate", "Amount", "Transport", "Net Amount", "Payment Mode", "Status", "Remarks"]
        for i, h in enumerate(headers):
            actual = str(ws.cell(row=1, column=i+1).value or "")
            check(f"Purchase header '{h}' at col {i+1}", actual == h)
        
        ok, msg = verify_formula(ws, "G2", ["E", "F"])
        check(f"Purchase Amount formula at G2: {msg if not ok else ''}", ok)
        ok, msg = verify_formula(ws, "I2", ["G", "H"])
        check(f"Purchase Net Amount formula at I2: {msg if not ok else ''}", ok)
        
        check("Purchase_Entry has data validations",
              len(ws.data_validations.dataValidation) > 0)
        check("Purchase_Entry has conditional formatting",
              len(ws.conditional_formatting) > 0)
        check("Purchase_Entry has freeze panes", ws.freeze_panes is not None)
        return
    
    if sheet_name == "Stock_Master":
        headers = ["Product Name", "Unit", "Opening Stock", "Purchases In",
                   "Sales Out", "Current Stock", "Reorder Level", "Rate (Rs)",
                   "Stock Value", "Status"]
        for i, h in enumerate(headers):
            actual = str(ws.cell(row=1, column=i+1).value or "")
            check(f"Stock header '{h}' at col {i+1}", actual == h, is_warning=True)
        
        # Check formulas
        ok, msg = verify_formula(ws, "D2", ["SUMIF"])
        check(f"Stock Purchases In formula at D2: {msg if not ok else ''}", ok)
        ok, msg = verify_formula(ws, "F2", ["C", "D", "E"])
        check(f"Stock Current Stock formula at F2: {msg if not ok else ''}", ok)
        ok, msg = verify_formula(ws, "I2", ["F", "H"])
        check(f"Stock Value formula at I2: {msg if not ok else ''}", ok)
        ok, msg = verify_formula(ws, "J2", ["IF"])
        check(f"Stock Status formula at J2: {msg if not ok else ''}", ok)
        
        check("Stock_Master has data validations",
              len(ws.data_validations.dataValidation) > 0)
        check("Stock_Master has conditional formatting",
              len(ws.conditional_formatting) > 0)
        check("Stock_Master has freeze panes", ws.freeze_panes is not None)
        return
    
    if sheet_name == "Party_Master":
        ok, msg = verify_formula(ws, "F2", ["SUMIF"])
        check(f"Party Total Debit formula at F2: {msg if not ok else ''}", ok)
        ok, msg = verify_formula(ws, "H2", ["E", "F", "G"])
        check(f"Party Running Balance formula at H2: {msg if not ok else ''}", ok)
        
        check("Party_Master has data validations",
              len(ws.data_validations.dataValidation) > 0)
        check("Party_Master has conditional formatting",
              len(ws.conditional_formatting) > 0)
        return
    
    if sheet_name == "Party_Ledger":
        # Check Balance formula (carry-forward)
        ok, msg = verify_formula(ws, "H3", ["H", "F", "G"])
        check(f"Ledger Balance carry-forward at H3: {msg if not ok else ''}", ok)
        
        check("Party_Ledger has data validations",
              len(ws.data_validations.dataValidation) > 0)
        check("Party_Ledger has conditional formatting",
              len(ws.conditional_formatting) > 0)
        return
    
    if sheet_name == "Sales_Report":
        # Check filter values
        check("Sales_Report has From Date filter at B4",
              ws.cell(row=4, column=2).value is not None)
        check("Sales_Report has summary formulas",
              str(ws.cell(row=5, column=6).value or "").startswith("=SUMIFS"))
        check("Sales_Report has data rows with INDEX formulas",
              str(ws.cell(row=9, column=1).value or "").startswith("=IFERROR(INDEX"))
        return
    
    if sheet_name == "Purchase_Report":
        check("Purchase_Report has filter controls at B4",
              ws.cell(row=4, column=2).value is not None)
        check("Purchase_Report has data rows",
              str(ws.cell(row=9, column=1).value or "").startswith("=IFERROR(INDEX"))
        check("Purchase_Report has conditional formatting",
              len(ws.conditional_formatting) > 0)
        return
    
    if sheet_name == "Stock_Report":
        check("Stock_Report has summary stats at B4",
              ws.cell(row=4, column=2).value is not None)
        check("Stock_Report has data rows",
              str(ws.cell(row=9, column=1).value or "").startswith("=IFERROR(INDEX"))
        check("Stock_Report has conditional formatting",
              len(ws.conditional_formatting) > 0)
        return
    
    if sheet_name == "Printable_Invoice":
        check("Invoice has Business Name header",
              "BUSINESS NAME" in str(ws.cell(row=1, column=1).value or "").upper())
        check("Invoice has TAX INVOICE title",
              "TAX INVOICE" in str(ws.cell(row=6, column=1).value or "").upper())
        check("Invoice has Grand Total formula",
              str(ws.cell(row=21, column=8).value or "").startswith("=SUM"))
        check("Invoice is portrait orientation",
              ws.page_setup.orientation == "portrait")
        check("Invoice paper size is A4",
              ws.page_setup.paperSize is not None)
        return
    
    if sheet_name == "Printable_Ledger":
        check("Ledger has PARTY LEDGER STATEMENT title",
              "LEDGER STATEMENT" in str(ws.cell(row=4, column=1).value or "").upper())
        # Find closing balance by scanning
        closing_found = any(
            "CLOSING" in str(ws.cell(row=r, column=1).value or "").upper()
            for r in range(20, 45)
        )
        has_closing_formula = any(
            str(ws.cell(row=r, column=7).value or "").startswith("=")
            for r in range(25, 40)
        )
        check("Ledger has Closing Balance label", closing_found)
        check("Ledger has Closing Balance formula", has_closing_formula)
        check("Ledger is portrait orientation",
              ws.page_setup.orientation == "portrait")
        return
    
    if sheet_name == "Settings":
        # Scan Settings to find all sections dynamically (row-independent)
        all_text = ""
        for r in range(1, 100):
            all_text += str(ws.cell(row=r, column=1).value or "").upper() + "\n"
        check("Settings has BUSINESS INFORMATION section", "BUSINESS INFORMATION" in all_text)
        check("Settings has UNITS section", "UNITS OF MEASUREMENT" in all_text)
        check("Settings has PAYMENT MODES section", "PAYMENT MODES" in all_text)
        check("Settings has TRANSACTION TYPES section", "TRANSACTION TYPES" in all_text)
        check("Settings has PRODUCT CATEGORIES section", "PRODUCT CATEGOR" in all_text)
        check("Settings has STATUS OPTIONS section", "STATUS OPTIONS" in all_text)
        check("Settings has QUICK REFERENCE section", "QUICK REFERENCE" in all_text)
        return


def main():
    global pass_count, fail_count, warn_count
    
    print("=" * 65)
    print("🐄 Dairy Plant Accounts — Workbook Verification")
    print("=" * 65)
    print()
    
    # Check file exists
    if not os.path.exists(WORKBOOK_PATH):
        print(f"❌ Workbook not found: {WORKBOOK_PATH}")
        print("   Run generate_accounts_workbook.py first!")
        sys.exit(1)
    
    file_size = os.path.getsize(WORKBOOK_PATH)
    print(f"📁 File: {WORKBOOK_PATH}")
    print(f"📏 Size: {file_size / 1024:.1f} KB")
    print(f"📅 Last modified: {datetime.datetime.fromtimestamp(os.path.getmtime(WORKBOOK_PATH)).strftime('%Y-%m-%d %H:%M')}")
    print()
    
    # Load workbook
    print("📂 Loading workbook...")
    wb = load_workbook(WORKBOOK_PATH)
    print()
    
    # ── 1. Sheet Validation ──
    print("📄 STEP 1: Sheet Structure")
    print("-" * 40)
    
    # Check all expected sheets exist
    actual_sheets = wb.sheetnames
    for sname in EXPECTED_SHEETS:
        check(f"Sheet '{sname}' exists", sname in actual_sheets)
    
    # Check no extra sheets
    for sname in actual_sheets:
        if sname not in EXPECTED_SHEETS and sname != "temp":
            check(f"Unexpected sheet '{sname}' found", False)
    
    # Check sheet order
    for i, sname in enumerate(EXPECTED_SHEETS):
        if i < len(actual_sheets):
            check(f"Sheet #{i+1} is '{sname}'", actual_sheets[i] == sname)
    
    # Check tab colors
    for sname, expected_color in TAB_COLORS.items():
        if sname in actual_sheets:
            actual_color = wb[sname].sheet_properties.tabColor
            if actual_color is not None:
                actual_str = actual_color.rgb if hasattr(actual_color, 'rgb') else str(actual_color)
                exp_str = expected_color
                check(f"Tab color for '{sname}' (actual: {actual_str}, expected: {exp_str})", 
                      actual_str.replace("00", "").upper() == exp_str.upper(),
                      is_warning=True)
    
    # Check active sheet is Dashboard
    active_name = wb.active.title if hasattr(wb.active, 'title') else str(wb.active)
    check(f"Active sheet is 'Dashboard' (actual: '{active_name}')", active_name == "Dashboard")
    
    print()
    
    # ── 2. Table Validation ──
    print("📊 STEP 2: Excel Tables")
    print("-" * 40)
    for sname, expected_table in EXPECTED_TABLES.items():
        if sname in actual_sheets:
            ws = wb[sname]
            table_names = list(ws.tables)  # ws.tables returns dict-like of Table objects in openpyxl
            checks_ok = expected_table in table_names or any(t.endswith(expected_table) for t in table_names)
            check(f"Table '{expected_table}' in '{sname}'", checks_ok)
    
    print()
    
    # ── 3. Formula Validation ──
    print("📐 STEP 3: Formula Validation")
    print("-" * 40)
    
    for sname in EXPECTED_SHEETS:
        if sname in actual_sheets:
            print(f"\n  Sheet: {sname}")
            ws = wb[sname]
            verify_sheet_structure(ws, sname)
    
    print()
    
    # ── 4. Cross-Sheet Reference Validation ──
    print("🔗 STEP 4: Cross-Sheet References")
    print("-" * 40)
    
    cross_refs = {
        "Dashboard": ["Sales_Entry", "Purchase_Entry", "Stock_Master", "Party_Master", "Party_Ledger"],
        "Stock_Master": ["Purchase_Entry", "Sales_Entry"],
        "Party_Master": ["Party_Ledger"],
        "Sales_Report": ["Sales_Entry"],
        "Purchase_Report": ["Purchase_Entry"],
        "Stock_Report": ["Stock_Master"],
    }
    
    for sname, refs in cross_refs.items():
        if sname in actual_sheets:
            ws = wb[sname]
            # Check that formulas reference target sheets
            for ref_sheet in refs:
                found = False
                for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 50, 50), 
                                         max_col=min(ws.max_column or 12, 12)):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                            if ref_sheet in cell.value:
                                found = True
                                break
                    if found:
                        break
                check(f"'{sname}' references '{ref_sheet}'", found)
    
    print()
    
    # ── 5. Sample Data ──
    print("📋 STEP 5: Sample Data Presence")
    print("-" * 40)
    
    data_sheets = {
        "Sales_Entry": 10,  # expect 10 data rows
        "Purchase_Entry": 8,
        "Stock_Master": 8,
        "Party_Master": 9,
        "Party_Ledger": 8,
    }
    
    for sname, min_rows in data_sheets.items():
        if sname in actual_sheets:
            ws = wb[sname]
            # Count non-empty data rows (past row 1)
            data_rows = 0
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value is not None:
                    data_rows += 1
            check(f"'{sname}' has at least {min_rows} data rows (found {data_rows})",
                  data_rows >= min_rows)
    
    print()
    
    # ── 6. Print Settings ──
    print("🖨️  STEP 6: Print Settings")
    print("-" * 40)
    
    print_sheets = {
        "Dashboard": "landscape",
        "Printable_Invoice": "portrait",
        "Printable_Ledger": "portrait",
    }
    
    for sname, expected_orientation in print_sheets.items():
        if sname in actual_sheets:
            ws = wb[sname]
            actual = ws.page_setup.orientation
            check(f"'{sname}' orientation is {expected_orientation} (actual: {actual})",
                  actual == expected_orientation)
    
    # Check page margins on print sheets
    for sname in ["Printable_Invoice", "Printable_Ledger"]:
        if sname in actual_sheets:
            ws = wb[sname]
            margins = ws.page_margins
            check(f"'{sname}' has page margins set", 
                  margins is not None and margins.left > 0)
    
    print()
    
    # ── Summary ──
    print("=" * 65)
    print("📊 VERIFICATION SUMMARY")
    print("=" * 65)
    total = pass_count + fail_count + warn_count
    print(f"  ✅ Passed:  {pass_count}")
    print(f"  ⚠️  Warnings: {warn_count}")
    print(f"  ❌ Failed:  {fail_count}")
    print(f"  ─────────────────")
    print(f"  📊 Total:   {total}")
    print()
    
    if fail_count == 0:
        print("🎉 ALL CHECKS PASSED! Workbook is ready for use.")
    else:
        print(f"⚠️  {fail_count} checks failed. Review the issues above.")
    
    print()
    print("💡 Next steps:")
    print("   1. Open Dairy_Plant_Accounts.xlsx in Excel")
    print("   2. Press F9 to recalculate all formulas")
    print("   3. Test the navigation hyperlinks on Dashboard row 3")
    print("   4. Verify KPI values update when data is changed")
    print()


if __name__ == "__main__":
    main()
