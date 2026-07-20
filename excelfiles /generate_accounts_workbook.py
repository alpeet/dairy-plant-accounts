#!/usr/bin/env python3
"""
Dairy Plant Accounts - Professional Excel Workbook Generator
==============================================================
Generates a complete accounting system workbook with two versions:
  1. Formula-only (.xlsx) — maximum compatibility with Windows 7 + macOS
  2. VBA-enhanced (.xlsm) — with PDF export buttons and automation

Usage:
  python3 generate_accounts_workbook.py

Output:
  Dairy_Plant_Accounts.xlsx   (Formula-only version)
  Dairy_Plant_Accounts_VBA.xlsm (VBA-enhanced version — needs VBA file imported)

Author: Freebuff AI / Buffy
Compatibility: Excel 2007+ (Windows 7), Excel 2011+ (macOS), LibreOffice
"""

import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.page import PageMargins

# ── CONSTANTS ──────────────────────────────────────────────────────────────
BUSINESS_NAME = "[Your Dairy Plant Name]"
VERSION = "1.0"
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_XLSX = os.path.join(OUTPUT_DIR, "Dairy_Plant_Accounts.xlsx")

TODAY = datetime.date.today()

# ── COLOR PALETTE (Professional Accounting Theme) ─────────────────────────
DARK_BLUE = "1B3A5C"
MED_BLUE = "2C5F8A"
LIGHT_BLUE = "D6E8F7"
ACCENT_GREEN = "27AE60"
ACCENT_RED = "E74C3C"
ACCENT_ORANGE = "F39C12"
BG_WHITE = "FFFFFF"
BG_LIGHT_GRAY = "F8F9FA"
BG_MED_GRAY = "E9ECEF"
TEXT_DARK = "2C3E50"
TEXT_MED = "6C757D"
BORDER_COLOR = "DEE2E6"

# Fills
HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
GREEN_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
RED_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
WHITE_FILL = PatternFill(start_color=BG_WHITE, end_color=BG_WHITE, fill_type="solid")
CARD_HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
KPI_FILL_1 = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
KPI_FILL_2 = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
KPI_FILL_3 = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
KPI_FILL_4 = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")

# Fonts
TITLE_FONT = Font(name="Calibri", bold=True, size=18, color=DARK_BLUE)
SUBTITLE_FONT = Font(name="Calibri", italic=True, size=10, color=TEXT_MED)
HEADER_FONT = Font(name="Calibri", bold=True, color=BG_WHITE, size=11)
SUBHEADER_FONT = Font(name="Calibri", bold=True, color=BG_WHITE, size=10)
KPI_VALUE_FONT = Font(name="Calibri", bold=True, size=22, color=DARK_BLUE)
KPI_LABEL_FONT = Font(name="Calibri", bold=True, size=9, color=TEXT_MED)
KPI_NUMBER_FONT = Font(name="Calibri", bold=True, size=18, color=MED_BLUE)
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
BOLD_12_FONT = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
LINK_FONT = Font(name="Calibri", size=10, color=MED_BLUE, underline="single")
SMALL_FONT = Font(name="Calibri", italic=True, size=8, color=TEXT_MED)
WHITE_BOLD_FONT = Font(name="Calibri", bold=True, size=10, color=BG_WHITE)

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)
BOTTOM_BORDER = Border(
    bottom=Side(style="medium", color=DARK_BLUE),
)
SECTION_BORDER = Border(
    left=Side(style="medium", color=DARK_BLUE),
    right=Side(style="medium", color=DARK_BLUE),
    top=Side(style="medium", color=DARK_BLUE),
    bottom=Side(style="medium", color=DARK_BLUE),
)
KPI_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="medium", color=MED_BLUE),
    bottom=Side(style="thin", color=BORDER_COLOR),
)
NO_BORDER = Border()

# Number formats
FMT_DATE = "DD-MMM-YYYY"
FMT_DATE_SHORT = "DD-MMM-YY"
FMT_NUM = '#,##0.00'
FMT_NUM_INT = '#,##0'
FMT_NUM_0DP = '#,##0'
FMT_NUM_PCT = '0.0%'
FMT_NUM_CR = '#,##0.00_ ;(#,##0.00)'

# ── HELPER FUNCTIONS ───────────────────────────────────────────────────────

def style_kpi_card(ws, row, col_start, col_end, label, formula, fmt=FMT_NUM, icon=""):
    """Create a styled KPI card spanning columns col_start:col_end."""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = f"{icon} {label}"
    cell.font = Font(name="Calibri", bold=True, size=9, color=TEXT_MED)
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    cell.fill = WHITE_FILL
    # Value row
    ws.merge_cells(start_row=row+1, start_column=col_start, end_row=row+1, end_column=col_end)
    val_cell = ws.cell(row=row+1, column=col_start)
    if formula:
        val_cell.value = formula
    val_cell.font = KPI_NUMBER_FONT
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    val_cell.number_format = fmt
    val_cell.fill = WHITE_FILL
    # Apply borders to all cells in range
    for c in range(col_start, col_end+1):
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row+1, column=c).border = KPI_BORDER


def add_title(ws, title, subtitle=None, row=1):
    """Add workbook title."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row=row, column=1, value=f"🐄 {BUSINESS_NAME}").font = TITLE_FONT
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    if subtitle:
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=12)
        ws.cell(row=row+1, column=1, value=subtitle).font = SUBTITLE_FONT
    return row + 3


def add_section_header(ws, row, text, col_start=1, col_end=12):
    """Add a section header with styling."""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    return row + 1


def apply_header_style(ws, headers, row=1):
    """Apply header styling with dark blue fill."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    last_col = get_column_letter(len(headers))
    # NOTE: AutoFilter NOT set here — Excel Tables have their own built-in
    # filter. Setting a sheet-level AutoFilter on a Table range causes
    # "Repair" errors in Excel. Add auto_filter manually on non-Table sheets.
    ws.freeze_panes = ws.cell(row=row+1, column=1).coordinate


def set_col_widths(ws, widths):
    """Set column widths from a list of (col_letter_or_num, width) pairs."""
    for col, w in widths:
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w


def add_excel_table(ws, headers, data, table_name, start_row=1, col_widths=None):
    """Create an Excel table with headers, data, and formatting."""
    row_count = len(data)
    col_count = len(headers)

    # Write header
    apply_header_style(ws, headers, row=start_row)

    # Write data
    for row_offset, row_data in enumerate(data):
        r = start_row + 1 + row_offset
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if isinstance(value, datetime.date):
                cell.number_format = FMT_DATE
            if isinstance(value, float):
                cell.number_format = FMT_NUM
        # Alternate row shading
        if row_offset % 2 == 1:
            for c in range(1, col_count + 1):
                ws.cell(row=r, column=c).fill = ALT_ROW_FILL

    # Create formal Excel Table
    if row_count > 0:
        end_row = start_row + row_count
        end_col = get_column_letter(col_count)
        ref = f"A{start_row}:{end_col}{end_row}"
        table = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    if col_widths:
        set_col_widths(ws, col_widths)


def add_data_validation(ws, col_letter, start_row, end_row, formula_list, title, error_msg):
    """Add dropdown data validation to a range."""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(formula_list)}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle=title,
        error=error_msg,
        showInputMessage=True,
        promptTitle=title,
        prompt=f"Select {title.lower()} from the list",
    )
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")
    return dv


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════

def build_dashboard(ws):
    ws.title = "Dashboard"

    # Title
    add_title(ws, BUSINESS_NAME, f"📊 Accounts Dashboard — Generated {TODAY.strftime('%d-%b-%Y')} | Version {VERSION}", row=1)
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 18

    # ── Row 3: Navigation bar with hyperlinks to all sheets ──
    nav_row = 3
    nav_sheets = [
        ("Daybook", "📒 Daybook"),
        ("Sales_Entry", "📝 Sales"),
        ("Purchase_Entry", "📦 Purchases"),
        ("Cash_Collection", "💵 Collections"),
        ("Stock_Master", "📋 Stock"),
        ("Party_Master", "👥 Parties"),
        ("Party_Ledger", "📊 Ledger"),
        ("Sales_Report", "📈 Sales Rpt"),
        ("Purchase_Report", "📉 Purchase Rpt"),
        ("Stock_Report", "🗃️ Stock Rpt"),
        ("Printable_Invoice", "🧾 Invoice"),
        ("Printable_Ledger", "📜 Statement"),
    ]
    for i, (sname, slabel) in enumerate(nav_sheets):
        c = i + 1
        cell = ws.cell(row=nav_row, column=c)
        cell.value = f'=HYPERLINK("#\'{sname}\'!A1","▶ {slabel}")'
        cell.font = Font(name="Calibri", size=9, color=DARK_BLUE, underline="single")
        cell.border = THIN_BORDER
        cell.fill = LIGHT_BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[nav_row].height = 22

    # ── Row 5-8: KPI Cards (shifted +1 for nav row) ──
    row = 5
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row+1].height = 50

    # Row 4-5: Top KPIs
    style_kpi_card(ws, row, 1, 3, "💰 MONTHLY SALES",
                   '=SUMIFS(Sales_Entry!J$2:J$5000,Sales_Entry!A$2:A$5000,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Sales_Entry!A$2:A$5000,"<="&EOMONTH(TODAY(),0))')
    style_kpi_card(ws, row, 4, 6, "📦 MONTHLY PURCHASES",
                   '=SUMIFS(Purchase_Entry!N$2:N$5000,Purchase_Entry!A$2:A$5000,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Purchase_Entry!A$2:A$5000,"<="&EOMONTH(TODAY(),0))')
    style_kpi_card(ws, row, 7, 9, "📋 STOCK VALUE",
                   '=SUMPRODUCT((Stock_Master!F$2:F$5000)*(Stock_Master!H$2:H$5000))')
    style_kpi_card(ws, row, 10, 12, "⚠️ LOW STOCK ITEMS",
                   '=COUNTIF(Stock_Master!J$2:J$5000,"Reorder")', FMT_NUM_INT, "⚠️")

    # Row 7-8: Second row KPIs (shifted +1)
    row = 7
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row+1].height = 50

    style_kpi_card(ws, row, 1, 3, "📤 OUTSTANDING RECEIVABLES",
                   '=SUMIFS(Sales_Entry!J$2:J$5000,Sales_Entry!K$2:K$5000,"Unpaid")+SUMIFS(Sales_Entry!J$2:J$5000,Sales_Entry!K$2:K$5000,"Partial")')
    style_kpi_card(ws, row, 4, 6, "📥 OUTSTANDING PAYABLES",
                   '=SUMIFS(Purchase_Entry!N$2:N$5000,Purchase_Entry!P$2:P$5000,"Unpaid")+SUMIFS(Purchase_Entry!N$2:N$5000,Purchase_Entry!P$2:P$5000,"Partial")')
    style_kpi_card(ws, row, 7, 9, "👥 TOTAL PARTIES",
                   '=COUNTA(Party_Master!A$2:A$5000)', FMT_NUM_INT)
    style_kpi_card(ws, row, 10, 12, "💰 DAYBOOK BALANCE",
                   '=IFERROR(INDEX(Daybook!G$2:G$5000,COUNTA(Daybook!A$2:A$5000)),0)', FMT_NUM)

    # ── Monthly Trend Table (shifted +1 for nav row) ──
    row = 11
    add_section_header(ws, row, "📈 MONTHLY TREND (Sales vs Purchases)")
    row += 1

    trend_headers = ["Month", "Sales", "Purchases", "Net Profit/Loss", "Invoice Count", "Purchase Count"]
    apply_header_style(ws, trend_headers, row=row)
    ws.row_dimensions[row].height = 22

    for i in range(6):
        r = row + 1 + i
        m = TODAY.month - i
        y = TODAY.year
        if m <= 0:
            m += 12
            y -= 1
        first_d = datetime.date(y, m, 1)
        if m == 12:
            last_d = datetime.date(y+1, 1, 1) - datetime.timedelta(days=1)
        else:
            last_d = datetime.date(y, m+1, 1) - datetime.timedelta(days=1)

        ws.cell(row=r, column=1, value=first_d.strftime("%b %Y")).font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER

        fy, fm, fd = first_d.year, first_d.month, first_d.day
        ly, lm, ld = last_d.year, last_d.month, last_d.day

        # Sales
        ws.cell(row=r, column=2).value = f'=SUMIFS(Sales_Entry!J$2:J$5000,Sales_Entry!A$2:A$5000,">="&DATE({fy},{fm},{fd}),Sales_Entry!A$2:A$5000,"<="&DATE({ly},{lm},{ld}))'
        ws.cell(row=r, column=2).number_format = FMT_NUM
        ws.cell(row=r, column=2).border = THIN_BORDER

        # Purchases (Net Amount is now column N)
        ws.cell(row=r, column=3).value = f'=SUMIFS(Purchase_Entry!N$2:N$5000,Purchase_Entry!A$2:A$5000,">="&DATE({fy},{fm},{fd}),Purchase_Entry!A$2:A$5000,"<="&DATE({ly},{lm},{ld}))'
        ws.cell(row=r, column=3).number_format = FMT_NUM
        ws.cell(row=r, column=3).border = THIN_BORDER

        # Net = Sales - Purchases
        ws.cell(row=r, column=4).value = f'=B{r}-C{r}'
        ws.cell(row=r, column=4).number_format = FMT_NUM
        ws.cell(row=r, column=4).font = BOLD_FONT
        ws.cell(row=r, column=4).border = THIN_BORDER

        # Invoice Count
        ws.cell(row=r, column=5).value = f'=COUNTIFS(Sales_Entry!A$2:A$5000,">="&DATE({fy},{fm},{fd}),Sales_Entry!A$2:A$5000,"<="&DATE({ly},{lm},{ld}))'
        ws.cell(row=r, column=5).number_format = FMT_NUM_INT
        ws.cell(row=r, column=5).border = THIN_BORDER

        # Purchase Count
        ws.cell(row=r, column=6).value = f'=COUNTIFS(Purchase_Entry!A$2:A$5000,">="&DATE({fy},{fm},{fd}),Purchase_Entry!A$2:A$5000,"<="&DATE({ly},{lm},{ld}))'
        ws.cell(row=r, column=6).number_format = FMT_NUM_INT
        ws.cell(row=r, column=6).border = THIN_BORDER

        # Alternating rows
        if i % 2 == 1:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = ALT_ROW_FILL

    # ── Party Balance Summary ──
    row = row + 2 + 6  # after trend table
    add_section_header(ws, row, "🏢 PARTY BALANCE SUMMARY")
    row += 1

    party_headers = ["Party Name", "Type", "Opening Balance", "Total Debit", "Total Credit", "Current Balance", "Status"]
    apply_header_style(ws, party_headers, row=row)
    # Add AutoFilter for quick party filtering from Dashboard
    ws.auto_filter.ref = f"A{row}:G{row+8}"
    # We'll put data-defined rows that users fill with party names
    for i in range(8):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=f"[Enter Party {i+1}]").font = Font(name="Calibri", italic=True, size=9, color=TEXT_MED)
        ws.cell(row=r, column=1).border = THIN_BORDER
        # Type from Party Master (column B = index 1 in VLOOKUP range B:G? Actually B$2:I$5000 range, index 1 = col B)
        ws.cell(row=r, column=2).value = f'=IFERROR(VLOOKUP(A{r},Party_Master!B$2:I$5000,1,FALSE),"")'
        ws.cell(row=r, column=2).border = THIN_BORDER
        # Opening Balance (column E = index 4 in VLOOKUP range B:I)
        ws.cell(row=r, column=3).value = f'=IFERROR(VLOOKUP(A{r},Party_Master!B$2:I$5000,4,FALSE),0)'
        ws.cell(row=r, column=3).number_format = FMT_NUM
        ws.cell(row=r, column=3).border = THIN_BORDER
        # Total Debit from Ledger
        ws.cell(row=r, column=4).value = f'=SUMIF(Party_Ledger!B$2:B$5000,A{r},Party_Ledger!F$2:F$5000)'
        ws.cell(row=r, column=4).number_format = FMT_NUM
        ws.cell(row=r, column=4).border = THIN_BORDER
        # Total Credit from Ledger
        ws.cell(row=r, column=5).value = f'=SUMIF(Party_Ledger!B$2:B$5000,A{r},Party_Ledger!G$2:G$5000)'
        ws.cell(row=r, column=5).number_format = FMT_NUM
        ws.cell(row=r, column=5).border = THIN_BORDER
        # Current Balance = Opening + Debit - Credit
        ws.cell(row=r, column=6).value = f'=C{r}+D{r}-E{r}'
        ws.cell(row=r, column=6).number_format = FMT_NUM
        ws.cell(row=r, column=6).font = BOLD_FONT
        ws.cell(row=r, column=6).border = THIN_BORDER
        # Status
        ws.cell(row=r, column=7).value = f'=IF(F{r}>0,"Receivable",IF(F{r}<0,"Payable","Settled"))'
        ws.cell(row=r, column=7).font = BOLD_FONT
        ws.cell(row=r, column=7).border = THIN_BORDER

    # ── Low Stock Alerts (using INDEX to Stock_Master rows) ──
    row = row + 10
    add_section_header(ws, row, "⚠️ LOW STOCK ALERTS")
    row += 1

    alert_headers = ["Product", "Current Stock", "Reorder Level", "Status"]
    apply_header_style(ws, alert_headers, row=row)
    for i in range(8):
        r = row + 1 + i
        src_row = i + 2
        ws.cell(row=r, column=1, value=f'=IFERROR(INDEX(Stock_Master!A$2:A$5000,{src_row}),"")').font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=f'=IF(A{r}="","",INDEX(Stock_Master!F$2:F$5000,{src_row}))').number_format = FMT_NUM_INT
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=3, value=f'=IF(A{r}="","",INDEX(Stock_Master!G$2:G$5000,{src_row}))').number_format = FMT_NUM_INT
        ws.cell(row=r, column=3).border = THIN_BORDER
        ws.cell(row=r, column=4, value=f'=IF(A{r}="","",INDEX(Stock_Master!J$2:J$5000,{src_row}))').font = BOLD_FONT
        ws.cell(row=r, column=4).border = THIN_BORDER

    # ── Recent Transactions ──
    row = row + 10
    add_section_header(ws, row, "🕐 RECENT TRANSACTIONS")
    row += 1

    recent_headers = ["Date", "Type", "Party", "Invoice/Bill", "Amount", "Status"]
    apply_header_style(ws, recent_headers, row=row)

    for i in range(5):
        r = row + 1 + i
        src_row = i + 2
        ws.cell(row=r, column=1, value=f'=IFERROR(INDEX(Sales_Entry!A$2:A$5000,{src_row}),"")').border = THIN_BORDER
        ws.cell(row=r, column=1).number_format = FMT_DATE_SHORT
        ws.cell(row=r, column=2, value='"Sale"').font = Font(name="Calibri", size=10, color=ACCENT_GREEN)
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=3, value=f'=IFERROR(INDEX(Sales_Entry!C$2:C$5000,{src_row}),"")').border = THIN_BORDER
        ws.cell(row=r, column=4, value=f'=IFERROR(INDEX(Sales_Entry!B$2:B$5000,{src_row}),"")').border = THIN_BORDER
        ws.cell(row=r, column=5, value=f'=IFERROR(INDEX(Sales_Entry!J$2:J$5000,{src_row}),"")').number_format = FMT_NUM
        ws.cell(row=r, column=5).border = THIN_BORDER
        ws.cell(row=r, column=6, value=f'=IFERROR(INDEX(Sales_Entry!K$2:K$5000,{src_row}),"")').border = THIN_BORDER

    # Column widths
    set_col_widths(ws, [(1, 22), (2, 18), (3, 18), (4, 18), (5, 18), (6, 18), (7, 18), (8, 16), (9, 16), (10, 16), (11, 16), (12, 16)])

    # Print settings
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

    # Conditional formatting for Net Profit/Loss (negative = red)
    ws.conditional_formatting.add(
        f"D{row+1}:D{row+6}",
        CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL, font=Font(name="Calibri", bold=True, color=ACCENT_RED, size=10))
    )


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2: SALES ENTRY
# ═══════════════════════════════════════════════════════════════════════════

def build_sales_entry(ws):
    ws.title = "Sales_Entry"

    headers = [
        "Date", "Invoice No", "Party Name", "Product", "Quantity",
        "Rate", "Amount", "Discount %", "Net Amount", "Payment Mode",
        "Status", "Remarks"
    ]

    # Sample data
    sample_parties = ["Sharma Dairy Shop", "Bhatbhateni Supermarket", "Patan Co-op", "Hotel Annapurna", "Khadgi Sweets"]
    sample_products = ["Raw Milk", "Cow Milk", "Curd (Dahi)", "Ghee", "Paneer"]
    sample_modes = ["Cash", "Credit", "Bank Transfer", "UPI", "Cheque"]
    sample_statuses = ["Paid", "Unpaid", "Partial"]

    data = []
    for i in range(10):
        d = TODAY - datetime.timedelta(days=i)
        inv = f"INV-{1001 + i}"
        p = sample_parties[i % 5]
        pr = sample_products[i % 5]
        qty = round(5 + (i * 2.5) + (i * 0.3), 2)
        rates = {"Raw Milk": 60, "Cow Milk": 65, "Curd (Dahi)": 100, "Ghee": 650, "Paneer": 350}
        rate = rates[pr]
        amt = round(qty * rate, 2)
        disc_pct = 0 if i % 4 == 0 else 2.5
        net = round(amt - (amt * disc_pct / 100), 2)
        mode = sample_modes[i % 5]
        status = sample_statuses[i % 3]
        data.append([d, inv, p, pr, qty, rate, amt, disc_pct, net, mode, status, ""])

    add_excel_table(ws, headers, data, "SalesTable", start_row=1,
                    col_widths=[(1, 14), (2, 14), (3, 24), (4, 20), (5, 10), (6, 10), (7, 12), (8, 10), (9, 14), (10, 14), (11, 10), (12, 22)])

    # Amount formula (column G)
    for r in range(2, 12):
        ws.cell(row=r, column=7).value = f'=E{r}*F{r}'
        ws.cell(row=r, column=7).number_format = FMT_NUM

    # Net Amount formula (column I)
    for r in range(2, 12):
        ws.cell(row=r, column=9).value = f'=G{r}-(G{r}*H{r}/100)'
        ws.cell(row=r, column=9).number_format = FMT_NUM

    # Data validations
    add_data_validation(ws, "C", 2, 5000, sample_parties, "Party Name", "Select a customer/party")
    add_data_validation(ws, "D", 2, 5000, sample_products, "Product", "Select a product")
    add_data_validation(ws, "J", 2, 5000, sample_modes, "Payment Mode", "Select payment mode")
    add_data_validation(ws, "K", 2, 5000, sample_statuses, "Status", "Select payment status")

    # Conditional formatting: highlight unpaid rows
    ws.conditional_formatting.add(
        f"A2:L5000",
        FormulaRule(formula=['$K2="Unpaid"'], fill=RED_FILL)
    )
    ws.conditional_formatting.add(
        f"A2:L5000",
        FormulaRule(formula=['$K2="Partial"'], fill=YELLOW_FILL)
    )
    ws.conditional_formatting.add(
        f"A2:L5000",
        FormulaRule(formula=['$K2="Paid"'], fill=GREEN_FILL)
    )


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3: PURCHASE ENTRY
# ═══════════════════════════════════════════════════════════════════════════

def build_purchase_entry(ws):
    ws.title = "Purchase_Entry"

    # ── REVISED: Milk Purchase Pricing Logic ──
    # Rate = FAT% * 7.15 + SNF% * 4.55 + ExtraPerUnit  (if RateType=FORMULA)
    #   or = Fixed Rate Per Unit                         (if RateType=FIXED)
    # Amount = Quantity * Rate
    # Net Amount = Amount + Transport

    headers = [
        "Date", "Bill No", "Supplier Name", "Product", "FAT %",
        "SNF %", "Extra Per Unit", "Rate Type", "Fixed Rate Per Unit",
        "Rate Per Unit", "Quantity", "Amount", "Transport",
        "Net Amount", "Payment Mode", "Status", "Remarks"
    ]

    suppliers = ["Gurung Dairy Farm", "Fresh Valley Suppliers", "Patan Co-op", "Pokhara Ghee House", "Dairy Cooperative"]
    products = ["Raw Milk", "Cow Milk", "Buffalo Milk", "Mixed Milk", "Ghee"]
    modes = ["Cash", "Credit", "Bank Transfer", "Cheque", "UPI"]
    statuses = ["Paid", "Unpaid", "Partial"]
    rate_types = ["FORMULA", "FIXED"]

    data = []
    for i in range(8):
        d = TODAY - datetime.timedelta(days=i)
        bill = f"BILL-{5001 + i}"
        s = suppliers[i % 5]
        pr = "Raw Milk" if i < 4 else products[i % 5]
        # Milk pricing data
        fat_pct = 3.5 + (i * 0.1) if pr in ["Raw Milk", "Cow Milk", "Buffalo Milk", "Mixed Milk"] else 0
        snf_pct = 8.5 + (i * 0.05) if pr in ["Raw Milk", "Cow Milk", "Buffalo Milk", "Mixed Milk"] else 0
        extra = 2.0 if i % 3 == 0 else 1.0 if i % 3 == 1 else 0.0
        rate_type = "FORMULA" if i < 6 else "FIXED"
        fixed_rate = 0.0 if rate_type == "FORMULA" else 50.0 + (i * 2)
        # Calculate rate per unit
        if rate_type == "FORMULA":
            rate_per_unit = round((fat_pct * 7.15) + (snf_pct * 4.55) + extra, 2)
        else:
            rate_per_unit = fixed_rate
        qty = round(50 + (i * 10) + (i * 2.5), 2)
        amt = round(qty * rate_per_unit, 2)
        transport = round(300 + i * 40, 2)
        net = round(amt + transport, 2)
        mode = modes[i % 5]
        status = statuses[i % 3]
        data.append([d, bill, s, pr, fat_pct, snf_pct, extra, rate_type, fixed_rate,
                     rate_per_unit, qty, amt, transport, net, mode, status, ""])

    add_excel_table(ws, headers, data, "PurchaseTable", start_row=1,
                    col_widths=[(1, 14), (2, 14), (3, 24), (4, 18), (5, 8), (6, 8), (7, 12),
                                (8, 12), (9, 14), (10, 14), (11, 10), (12, 14), (13, 12),
                                (14, 14), (15, 14), (16, 10), (17, 22)])

    # ── RATE PER UNIT FORMULA (Col J = column 10) ──
    # If RateType = FORMULA: (FAT% * 7.15) + (SNF% * 4.55) + ExtraPerUnit
    # If RateType = FIXED: Fixed Rate Per Unit
    for r in range(2, 10):
        ws.cell(row=r, column=10).value = (
            f'=IF(H{r}="FORMULA",'
            f' IF(E{r}="",0,E{r})*7.15 + IF(F{r}="",0,F{r})*4.55 + IF(G{r}="",0,G{r}),'
            f' IF(H{r}="FIXED", IF(I{r}="",0,I{r}), 0))'
        )
        ws.cell(row=r, column=10).number_format = FMT_NUM

        # Amount = Quantity * Rate Per Unit (Col L = column 12)
        ws.cell(row=r, column=12).value = f'=IF(K{r}="",0,K{r})*J{r}'
        ws.cell(row=r, column=12).number_format = FMT_NUM

        # Net Amount = Amount + Transport (Col N = column 14)
        ws.cell(row=r, column=14).value = f'=L{r}+IF(M{r}="",0,M{r})'
        ws.cell(row=r, column=14).number_format = FMT_NUM

    # ── DATA VALIDATIONS ──
    add_data_validation(ws, "C", 2, 5000, suppliers, "Supplier", "Select a supplier")
    add_data_validation(ws, "D", 2, 5000, products, "Product", "Select a product")
    add_data_validation(ws, "H", 2, 5000, rate_types, "Rate Type", "Select FORMULA or FIXED")
    add_data_validation(ws, "O", 2, 5000, modes, "Payment Mode", "Select payment mode")
    add_data_validation(ws, "P", 2, 5000, statuses, "Status", "Select payment status")

    # ── CONDITIONAL FORMATTING ──
    ws.conditional_formatting.add("A2:Q5000", FormulaRule(formula=['$P2="Unpaid"'], fill=RED_FILL))
    ws.conditional_formatting.add("A2:Q5000", FormulaRule(formula=['$P2="Partial"'], fill=YELLOW_FILL))
    ws.conditional_formatting.add("A2:Q5000", FormulaRule(formula=['$P2="Paid"'], fill=GREEN_FILL))

    # ── PRICING INFO / NOTES BELOW TABLE ──
    info_row = 12
    ws.merge_cells(f"A{info_row}:Q{info_row}")
    ws.cell(row=info_row, column=1, value="📌 MILK PURCHASE PRICING RULES").font = BOLD_12_FONT

    notes = [
        "Rate Per Unit = (FAT % × 7.15) + (SNF % × 4.55) + Extra Per Unit   when Rate Type = \"FORMULA\"",
        "Rate Per Unit = Fixed Rate Per Unit                                 when Rate Type = \"FIXED\"",
        "Amount = Quantity × Rate Per Unit",
        "Net Amount = Amount + Transport Charges",
        "FAT Rate Multiplier: 7.15 per unit | SNF Rate Multiplier: 4.55 per unit",
        "Extra Per Unit is any additional contract amount (e.g., +2 per liter)",
        "💡 Select Rate Type = FIXED for simple flat-rate purchases (e.g., Rs. 50/liter)",
    ]
    for i, note in enumerate(notes):
        r = info_row + 1 + i
        ws.merge_cells(f"A{r}:Q{r}")
        ws.cell(row=r, column=1, value=note).font = Font(name="Calibri", italic=True, size=9, color=TEXT_MED)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4: STOCK MASTER
# ═══════════════════════════════════════════════════════════════════════════

def build_stock_master(ws):
    ws.title = "Stock_Master"

    headers = [
        "Product Name", "Unit", "Opening Stock", "Purchases In",
        "Sales Out", "Current Stock", "Reorder Level", "Rate (Rs)",
        "Stock Value", "Status"
    ]

    products = [
        ["Raw Milk", "Liter", 1000, 0, 0, 0, 200, 60, 0, ""],
        ["Cow Milk", "Liter", 500, 0, 0, 0, 100, 65, 0, ""],
        ["Curd (Dahi)", "Kg", 100, 0, 0, 0, 20, 100, 0, ""],
        ["Ghee", "Kg", 50, 0, 0, 0, 10, 650, 0, ""],
        ["Paneer", "Kg", 30, 0, 0, 0, 5, 350, 0, ""],
        ["Butter", "Kg", 20, 0, 0, 0, 5, 450, 0, ""],
        ["Buttermilk", "Liter", 60, 0, 0, 0, 15, 40, 0, ""],
        ["Khoya", "Kg", 15, 0, 0, 0, 5, 300, 0, ""],
    ]

    add_excel_table(ws, headers, products, "StockMasterTable", start_row=1,
                    col_widths=[(1, 22), (2, 10), (3, 14), (4, 14), (5, 12), (6, 14), (7, 14), (8, 12), (9, 14), (10, 14)])

    # Formulas
    for r in range(2, 2 + len(products)):
        # Purchases In (from Purchase_Entry, SUM by product; Quantity is now column K)
        ws.cell(row=r, column=4).value = f'=SUMIF(Purchase_Entry!D$2:D$5000,A{r},Purchase_Entry!K$2:K$5000)'
        ws.cell(row=r, column=4).number_format = FMT_NUM_INT
        # Sales Out (from Sales_Entry, SUM by product)
        ws.cell(row=r, column=5).value = f'=SUMIF(Sales_Entry!D$2:D$5000,A{r},Sales_Entry!E$2:E$5000)'
        ws.cell(row=r, column=5).number_format = FMT_NUM_INT
        # Current Stock = Opening + Purchases - Sales
        ws.cell(row=r, column=6).value = f'=C{r}+D{r}-E{r}'
        ws.cell(row=r, column=6).font = BOLD_FONT
        ws.cell(row=r, column=6).number_format = FMT_NUM_INT
        # Stock Value = Current Stock * Rate
        ws.cell(row=r, column=9).value = f'=F{r}*H{r}'
        ws.cell(row=r, column=9).number_format = FMT_NUM
        ws.cell(row=r, column=9).font = BOLD_FONT
        # Status
        ws.cell(row=r, column=10).value = f'=IF(F{r}<=G{r},"Reorder",IF(F{r}<=(G{r}*1.5),"Low Stock","In Stock"))'

    # Data validation: Unit
    add_data_validation(ws, "B", 2, 5000, ["Litre", "Kg", "Piece", "Packet", "Box"], "Unit", "Select unit")

    # Conditional formatting
    ws.conditional_formatting.add(
        "A2:J5000",
        FormulaRule(formula=['$J2="Reorder"'], fill=RED_FILL, font=Font(name="Calibri", bold=True, size=10))
    )
    ws.conditional_formatting.add(
        "A2:J5000",
        FormulaRule(formula=['$J2="Low Stock"'], fill=YELLOW_FILL)
    )
    ws.conditional_formatting.add(
        "A2:J5000",
        FormulaRule(formula=['$J2="In Stock"'], fill=GREEN_FILL)
    )
    ws.conditional_formatting.add(
        "F2:F5000",
        CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL)
    )


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5: PARTY MASTER
# ═══════════════════════════════════════════════════════════════════════════

def build_party_master(ws):
    ws.title = "Party_Master"

    headers = [
        "Party Name", "Type", "Phone", "Address", "Opening Balance",
        "Total Debit", "Total Credit", "Running Balance", "Status"
    ]

    parties = [
        ["Sharma Dairy Shop", "Customer", "9812345670", "New Road, Kathmandu", 15000, 0, 0, 0, ""],
        ["Bhatbhateni Supermarket", "Customer", "9812345672", "Durbar Marg, Kathmandu", 45000, 0, 0, 0, ""],
        ["Gurung Dairy Farm", "Supplier", "9852345673", "Budhanilkantha", 50000, 0, 0, 0, ""],
        ["Fresh Valley Suppliers", "Supplier", "9841234570", "Bhaktapur", 12000, 0, 0, 0, ""],
        ["Patan Co-op", "Both", "9852345675", "Patan, Lalitpur", 0, 0, 0, 0, ""],
        ["Hotel Annapurna", "Customer", "9811112233", "Thamel, Kathmandu", 0, 0, 0, 0, ""],
        ["Khadgi Sweets", "Customer", "9843322110", "Asan, Kathmandu", 8000, 0, 0, 0, ""],
        ["Pokhara Ghee House", "Supplier", "9856001122", "Pokhara", 25000, 0, 0, 0, ""],
        ["Dairy Cooperative", "Both", "9812345699", "Chitwan", 0, 0, 0, 0, ""],
    ]

    add_excel_table(ws, headers, parties, "PartyMasterTable", start_row=1,
                    col_widths=[(1, 24), (2, 12), (3, 16), (4, 26), (5, 16), (6, 14), (7, 14), (8, 16), (9, 14)])

    # Formulas
    for r in range(2, 2 + len(parties)):
        # Total Debit from Party_Ledger
        ws.cell(row=r, column=6).value = f'=SUMIF(Party_Ledger!B$2:B$5000,A{r},Party_Ledger!F$2:F$5000)'
        ws.cell(row=r, column=6).number_format = FMT_NUM
        # Total Credit from Party_Ledger
        ws.cell(row=r, column=7).value = f'=SUMIF(Party_Ledger!B$2:B$5000,A{r},Party_Ledger!G$2:G$5000)'
        ws.cell(row=r, column=7).number_format = FMT_NUM
        # Running Balance = Opening + Debit - Credit
        ws.cell(row=r, column=8).value = f'=E{r}+F{r}-G{r}'
        ws.cell(row=r, column=8).number_format = FMT_NUM
        ws.cell(row=r, column=8).font = BOLD_FONT
        # Status
        ws.cell(row=r, column=9).value = f'=IF(H{r}>0,"Outstanding",IF(H{r}<0,"Payable","Settled"))'

    # Data validations
    add_data_validation(ws, "B", 2, 5000, ["Customer", "Supplier", "Both"], "Party Type", "Select party type")

    # Conditional formatting
    ws.conditional_formatting.add("A2:J5000", FormulaRule(formula=['$H2>0'], fill=YELLOW_FILL))
    ws.conditional_formatting.add("A2:J5000", FormulaRule(formula=['$H2<0'], fill=RED_FILL))


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6: PARTY LEDGER
# ═══════════════════════════════════════════════════════════════════════════

def build_party_ledger(ws):
    ws.title = "Party_Ledger"

    headers = [
        "Date", "Party Name", "Transaction Type", "Reference No",
        "Description", "Debit", "Credit", "Balance", "Remarks"
    ]

    data = [
        [datetime.date(2026, 1, 1), "Sharma Dairy Shop", "Opening", "", "Opening Balance", 15000, 0, 15000, ""],
        [TODAY - datetime.timedelta(days=7), "Sharma Dairy Shop", "Sale", "INV-1001", "Sale of Milk & Products", 5800, 0, 20800, ""],
        [TODAY - datetime.timedelta(days=5), "Sharma Dairy Shop", "Receipt", "PAY-101", "Payment Received", 0, 10000, 10800, "Bank Transfer"],
        [TODAY - datetime.timedelta(days=3), "Sharma Dairy Shop", "Sale", "INV-1002", "Ghee Sale", 3250, 0, 14050, ""],
        [datetime.date(2026, 1, 1), "Gurung Dairy Farm", "Opening", "", "Opening Balance", 0, 50000, -50000, ""],
        [TODAY - datetime.timedelta(days=6), "Gurung Dairy Farm", "Purchase", "BILL-5001", "Milk Purchase", 0, 15300, -65300, "Transport included"],
        [TODAY - datetime.timedelta(days=4), "Gurung Dairy Farm", "Payment", "PAY-102", "Payment Made", 30000, 0, -35300, "Bank Transfer"],
        [TODAY - datetime.timedelta(days=2), "Gurung Dairy Farm", "Purchase", "BILL-5002", "Ghee Supply", 0, 8400, -43700, ""],
    ]

    add_excel_table(ws, headers, data, "PartyLedgerTable", start_row=1,
                    col_widths=[(1, 14), (2, 24), (3, 16), (4, 14), (5, 30), (6, 14), (7, 14), (8, 14), (9, 22)])

    # Balance formula with carry-forward
    for r in range(3, 2 + len(data)):
        ws.cell(row=r, column=8).value = f'=H{r-1}+F{r}-G{r}'
        ws.cell(row=r, column=8).number_format = FMT_NUM
        ws.cell(row=r, column=8).font = BOLD_FONT

    # Data validations
    add_data_validation(ws, "B", 2, 5000,
                        ["Sharma Dairy Shop", "Bhatbhateni Supermarket", "Gurung Dairy Farm", "Fresh Valley Suppliers",
                         "Patan Co-op", "Hotel Annapurna", "Khadgi Sweets", "Pokhara Ghee House", "Dairy Cooperative"],
                        "Party Name", "Select a party")
    add_data_validation(ws, "C", 2, 5000,
                        ["Sale", "Purchase", "Receipt", "Payment", "Opening", "Adjustment", "Credit Note", "Debit Note"],
                        "Transaction Type", "Select type")

    # Conditional formatting
    ws.conditional_formatting.add("A2:I5000", FormulaRule(formula=['$H2<0'], fill=RED_FILL))
    ws.conditional_formatting.add("A2:I5000", FormulaRule(formula=['$H2>0'], fill=GREEN_FILL))


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 7: SALES REPORT
# ═══════════════════════════════════════════════════════════════════════════

def build_sales_report(ws):
    ws.title = "Sales_Report"

    # Report controls
    ws.cell(row=1, column=1, value="🐄 SALES REPORT").font = TITLE_FONT
    ws.merge_cells("A1:K1")

    # Filters
    ws.cell(row=3, column=1, value="Filters:").font = BOLD_12_FONT
    ws.cell(row=4, column=1, value="From Date:").font = BOLD_FONT
    ws.cell(row=4, column=2, value=f'={TODAY.replace(day=1).strftime("%d-%b-%Y")}').font = NORMAL_FONT
    ws.cell(row=4, column=2).number_format = FMT_DATE
    ws.cell(row=5, column=1, value="To Date:").font = BOLD_FONT
    ws.cell(row=5, column=2, value=f'={TODAY.strftime("%d-%b-%Y")}').font = NORMAL_FONT
    ws.cell(row=5, column=2).number_format = FMT_DATE
    ws.cell(row=6, column=1, value="Party:").font = BOLD_FONT
    ws.cell(row=6, column=2, value="All").font = NORMAL_FONT

    # Summary row
    ws.cell(row=3, column=5, value="Summary:").font = BOLD_12_FONT
    ws.cell(row=4, column=5, value="Total Invoices:").font = BOLD_FONT
    ws.cell(row=4, column=6, value='=COUNTIFS(Sales_Entry!A$2:A$5000,">="&B4,Sales_Entry!A$2:A$5000,"<="&B5)').font = BOLD_FONT
    ws.cell(row=5, column=5, value="Total Amount:").font = BOLD_FONT
    ws.cell(row=5, column=6, value='=SUMIFS(Sales_Entry!J$2:J$5000,Sales_Entry!A$2:A$5000,">="&B4,Sales_Entry!A$2:A$5000,"<="&B5)').font = BOLD_FONT
    ws.cell(row=5, column=6).number_format = FMT_NUM
    ws.cell(row=6, column=5, value="Outstanding:").font = BOLD_FONT
    ws.cell(row=6, column=6, value='=SUMIFS(Sales_Entry!J$2:J$5000,Sales_Entry!A$2:A$5000,">="&B4,Sales_Entry!A$2:A$5000,"<="&B5,Sales_Entry!K$2:K$5000,"Unpaid")+SUMIFS(Sales_Entry!J$2:J$5000,Sales_Entry!A$2:A$5000,">="&B4,Sales_Entry!A$2:A$5000,"<="&B5,Sales_Entry!K$2:K$5000,"Partial")').font = Font(name="Calibri", bold=True, color=ACCENT_RED, size=10)
    ws.cell(row=6, column=6).number_format = FMT_NUM

    # Data section - simple row-by-row reference from Sales_Entry
    row = 8
    report_headers = ["Date", "Invoice No", "Party Name", "Product", "Quantity", "Rate", "Amount", "Discount %", "Net Amount", "Payment Mode", "Status"]
    apply_header_style(ws, report_headers, row=row)

    # Add AutoFilter on data rows only (not on totals row)
    ws.auto_filter.ref = f"A{row}:K{row+20}"

    # Simple direct row references - no CSE array formulas needed
    for i in range(20):
        r = row + 1 + i
        src_row = i + 2  # Row index in Sales_Entry
        ws.cell(row=r, column=1, value=f'=IFERROR(INDEX(Sales_Entry!A$2:A$5000,{src_row}),"")').font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).number_format = FMT_DATE_SHORT

        # Direct INDEX per column
        ws.cell(row=r, column=2, value=f'=IFERROR(INDEX(Sales_Entry!B$2:B$5000,{src_row}),"")').border = THIN_BORDER
        ws.cell(row=r, column=3, value=f'=IFERROR(INDEX(Sales_Entry!C$2:C$5000,{src_row}),"")').border = THIN_BORDER
        ws.cell(row=r, column=4, value=f'=IFERROR(INDEX(Sales_Entry!D$2:D$5000,{src_row}),"")').border = THIN_BORDER
        ws.cell(row=r, column=5, value=f'=IFERROR(INDEX(Sales_Entry!E$2:E$5000,{src_row}),"")').border = THIN_BORDER
        ws.cell(row=r, column=5).number_format = FMT_NUM
        ws.cell(row=r, column=6, value=f'=IFERROR(INDEX(Sales_Entry!F$2:F$5000,{src_row}),"")').border = THIN_BORDER
        ws.cell(row=r, column=6).number_format = FMT_NUM
        ws.cell(row=r, column=7, value=f'=IFERROR(INDEX(Sales_Entry!G$2:G$5000,{src_row}),"")').border = THIN_BORDER
        ws.cell(row=r, column=7).number_format = FMT_NUM
        ws.cell(row=r, column=8, value=f'=IFERROR(INDEX(Sales_Entry!H$2:H$5000,{src_row}),"")').border = THIN_BORDER
        ws.cell(row=r, column=9, value=f'=IFERROR(INDEX(Sales_Entry!I$2:I$5000,{src_row}),"")').border = THIN_BORDER
        ws.cell(row=r, column=9).number_format = FMT_NUM
        ws.cell(row=r, column=10, value=f'=IFERROR(INDEX(Sales_Entry!J$2:J$5000,{src_row}),"")').border = THIN_BORDER
        ws.cell(row=r, column=11, value=f'=IFERROR(INDEX(Sales_Entry!K$2:K$5000,{src_row}),"")').border = THIN_BORDER

    # Totals row
    tot_row = row + 21
    ws.cell(row=tot_row, column=1, value="TOTALS").font = BOLD_FONT
    ws.cell(row=tot_row, column=1).border = THIN_BORDER
    ws.cell(row=tot_row, column=7, value=f'=SUM(G{row+1}:G{tot_row-1})').font = BOLD_FONT
    ws.cell(row=tot_row, column=7).number_format = FMT_NUM
    ws.cell(row=tot_row, column=7).border = THIN_BORDER
    ws.cell(row=tot_row, column=9, value=f'=SUM(I{row+1}:I{tot_row-1})').font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
    ws.cell(row=tot_row, column=9).number_format = FMT_NUM
    ws.cell(row=tot_row, column=9).border = THIN_BORDER

    set_col_widths(ws, [(1, 14), (2, 14), (3, 24), (4, 20), (5, 10), (6, 10), (7, 14), (8, 12), (9, 14), (10, 14), (11, 12)])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 8: PURCHASE REPORT
# ═══════════════════════════════════════════════════════════════════════════

def build_purchase_report(ws):
    ws.title = "Purchase_Report"

    ws.cell(row=1, column=1, value="🐄 PURCHASE REPORT").font = TITLE_FONT
    ws.merge_cells("A1:K1")

    ws.cell(row=3, column=1, value="Filters:").font = BOLD_12_FONT
    ws.cell(row=4, column=1, value="From Date:").font = BOLD_FONT
    ws.cell(row=4, column=2, value=f'={TODAY.replace(day=1).strftime("%d-%b-%Y")}').number_format = FMT_DATE
    ws.cell(row=5, column=1, value="To Date:").font = BOLD_FONT
    ws.cell(row=5, column=2, value=f'={TODAY.strftime("%d-%b-%Y")}').number_format = FMT_DATE
    ws.cell(row=6, column=1, value="Supplier:").font = BOLD_FONT
    ws.cell(row=6, column=2, value="All").font = NORMAL_FONT

    ws.cell(row=3, column=5, value="Summary:").font = BOLD_12_FONT
    ws.cell(row=4, column=5, value="Total Bills:").font = BOLD_FONT
    ws.cell(row=4, column=6, value='=COUNTIFS(Purchase_Entry!A$2:A$5000,">="&B4,Purchase_Entry!A$2:A$5000,"<="&B5)')
    ws.cell(row=5, column=5, value="Total Amount:").font = BOLD_FONT
    ws.cell(row=5, column=6, value='=SUMIFS(Purchase_Entry!N$2:N$5000,Purchase_Entry!A$2:A$5000,">="&B4,Purchase_Entry!A$2:A$5000,"<="&B5)')
    ws.cell(row=5, column=6).number_format = FMT_NUM
    ws.cell(row=6, column=5, value="Payables:").font = BOLD_FONT
    ws.cell(row=6, column=6, value='=SUMIFS(Purchase_Entry!N$2:N$5000,Purchase_Entry!A$2:A$5000,">="&B4,Purchase_Entry!A$2:A$5000,"<="&B5,Purchase_Entry!P$2:P$5000,"Unpaid")+SUMIFS(Purchase_Entry!N$2:N$5000,Purchase_Entry!A$2:A$5000,">="&B4,Purchase_Entry!A$2:A$5000,"<="&B5,Purchase_Entry!P$2:P$5000,"Partial")')
    ws.cell(row=6, column=6).number_format = FMT_NUM
    ws.cell(row=6, column=6).font = Font(name="Calibri", bold=True, color=ACCENT_RED, size=10)

    row = 8
    report_headers = ["Date", "Bill No", "Supplier", "Product", "Quantity", "Rate", "Amount", "Transport", "Net Amount", "Payment Mode", "Status"]
    apply_header_style(ws, report_headers, row=row)

    # Add AutoFilter on data rows only (not on totals row)
    ws.auto_filter.ref = f"A{row}:K{row+20}"

    # Updated for new Purchase_Entry column layout:
    # A=Date, B=BillNo, C=Supplier, D=Product, E=FAT%, F=SNF%, G=Extra, H=RateType,
    # I=FixedRate, J=RatePerUnit, K=Quantity, L=Amount, M=Transport, N=NetAmount
    # O=PaymentMode, P=Status, Q=Remarks
    for i in range(20):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=f'=IFERROR(INDEX(Purchase_Entry!A$2:A$5000,{i+2}),"")').border = THIN_BORDER
        ws.cell(row=r, column=1).number_format = FMT_DATE_SHORT
        ws.cell(row=r, column=2, value=f'=IF(A{r}="","",INDEX(Purchase_Entry!B$2:B$5000,{i+2}))').border = THIN_BORDER
        ws.cell(row=r, column=3, value=f'=IF(A{r}="","",INDEX(Purchase_Entry!C$2:C$5000,{i+2}))').border = THIN_BORDER
        ws.cell(row=r, column=4, value=f'=IF(A{r}="","",INDEX(Purchase_Entry!D$2:D$5000,{i+2}))').border = THIN_BORDER
        ws.cell(row=r, column=5, value=f'=IF(A{r}="","",INDEX(Purchase_Entry!K$2:K$5000,{i+2}))').border = THIN_BORDER  # Quantity
        ws.cell(row=r, column=5).number_format = FMT_NUM
        ws.cell(row=r, column=6, value=f'=IF(A{r}="","",INDEX(Purchase_Entry!J$2:J$5000,{i+2}))').border = THIN_BORDER  # Rate Per Unit
        ws.cell(row=r, column=6).number_format = FMT_NUM
        ws.cell(row=r, column=7, value=f'=IF(A{r}="","",INDEX(Purchase_Entry!L$2:L$5000,{i+2}))').border = THIN_BORDER  # Amount
        ws.cell(row=r, column=7).number_format = FMT_NUM
        ws.cell(row=r, column=8, value=f'=IF(A{r}="","",INDEX(Purchase_Entry!M$2:M$5000,{i+2}))').border = THIN_BORDER  # Transport
        ws.cell(row=r, column=8).number_format = FMT_NUM
        ws.cell(row=r, column=9, value=f'=IF(A{r}="","",INDEX(Purchase_Entry!N$2:N$5000,{i+2}))').border = THIN_BORDER  # Net Amount
        ws.cell(row=r, column=9).number_format = FMT_NUM
        ws.cell(row=r, column=10, value=f'=IF(A{r}="","",INDEX(Purchase_Entry!O$2:O$5000,{i+2}))').border = THIN_BORDER  # Payment Mode
        ws.cell(row=r, column=11, value=f'=IF(A{r}="","",INDEX(Purchase_Entry!P$2:P$5000,{i+2}))').border = THIN_BORDER  # Status

    tot_row = row + 21
    ws.cell(row=tot_row, column=1, value="TOTALS").font = BOLD_FONT
    ws.cell(row=tot_row, column=1).border = THIN_BORDER
    ws.cell(row=tot_row, column=7, value=f'=SUM(G{row+1}:G{tot_row-1})').font = BOLD_FONT
    ws.cell(row=tot_row, column=7).number_format = FMT_NUM
    ws.cell(row=tot_row, column=7).border = THIN_BORDER
    ws.cell(row=tot_row, column=9, value=f'=SUM(I{row+1}:I{tot_row-1})').font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
    ws.cell(row=tot_row, column=9).number_format = FMT_NUM
    ws.cell(row=tot_row, column=9).border = THIN_BORDER
    ws.cell(row=tot_row, column=8, value=f'=SUM(H{row+1}:H{tot_row-1})').font = BOLD_FONT
    ws.cell(row=tot_row, column=8).number_format = FMT_NUM
    ws.cell(row=tot_row, column=8).border = THIN_BORDER

    set_col_widths(ws, [(1, 14), (2, 14), (3, 24), (4, 20), (5, 10), (6, 10), (7, 14), (8, 12), (9, 14), (10, 14), (11, 12)])

    # Conditional formatting for unpaid
    ws.conditional_formatting.add(f"A{row+1}:K{tot_row-1}", FormulaRule(formula=[f'$K{row+1}="Unpaid"'], fill=RED_FILL))
    ws.conditional_formatting.add(f"A{row+1}:K{tot_row-1}", FormulaRule(formula=[f'$K{row+1}="Partial"'], fill=YELLOW_FILL))


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 9: STOCK REPORT
# ═══════════════════════════════════════════════════════════════════════════

def build_stock_report(ws):
    ws.title = "Stock_Report"

    ws.cell(row=1, column=1, value="🐄 STOCK REPORT").font = TITLE_FONT
    ws.merge_cells("A1:H1")

    ws.cell(row=3, column=1, value="Stock Summary:").font = BOLD_12_FONT
    ws.cell(row=4, column=1, value="Total Products:").font = BOLD_FONT
    ws.cell(row=4, column=2, value='=COUNTA(Stock_Master!A$2:A$5000)')
    ws.cell(row=5, column=1, value="Total Stock Value:").font = BOLD_FONT
    ws.cell(row=5, column=2, value='=SUMPRODUCT((Stock_Master!F$2:F$5000)*(Stock_Master!H$2:H$5000))')
    ws.cell(row=5, column=2).number_format = FMT_NUM
    ws.cell(row=6, column=1, value="Low Stock Items:").font = BOLD_FONT
    ws.cell(row=6, column=2, value='=COUNTIF(Stock_Master!J$2:J$5000,"Reorder")')
    ws.cell(row=6, column=2).font = Font(name="Calibri", bold=True, color=ACCENT_RED, size=10)

    row = 8
    headers = ["Product", "Unit", "Opening", "Purchases", "Sales", "Current", "Reorder Level", "Stock Value", "Status"]
    apply_header_style(ws, headers, row=row)

    # Add AutoFilter on data rows only (not on totals row)
    ws.auto_filter.ref = f"A{row}:I{row+20}"

    # Link to Stock Master
    for i in range(20):
        r = row + 1 + i
        idx = i + 2
        ws.cell(row=r, column=1, value=f'=IFERROR(INDEX(Stock_Master!A$2:A$5000,{idx}),"")').border = THIN_BORDER
        ws.cell(row=r, column=2, value=f'=IF(A{r}="","",INDEX(Stock_Master!B$2:B$5000,{idx}))').border = THIN_BORDER
        ws.cell(row=r, column=3, value=f'=IF(A{r}="","",INDEX(Stock_Master!C$2:C$5000,{idx}))').border = THIN_BORDER
        ws.cell(row=r, column=4, value=f'=IF(A{r}="","",INDEX(Stock_Master!D$2:D$5000,{idx}))').border = THIN_BORDER
        ws.cell(row=r, column=5, value=f'=IF(A{r}="","",INDEX(Stock_Master!E$2:E$5000,{idx}))').border = THIN_BORDER
        ws.cell(row=r, column=6, value=f'=IF(A{r}="","",INDEX(Stock_Master!F$2:F$5000,{idx}))').font = BOLD_FONT
        ws.cell(row=r, column=6).border = THIN_BORDER
        ws.cell(row=r, column=7, value=f'=IF(A{r}="","",INDEX(Stock_Master!G$2:G$5000,{idx}))').border = THIN_BORDER
        ws.cell(row=r, column=8, value=f'=IF(A{r}="","",INDEX(Stock_Master!I$2:I$5000,{idx}))').border = THIN_BORDER
        ws.cell(row=r, column=8).number_format = FMT_NUM
        ws.cell(row=r, column=9, value=f'=IF(A{r}="","",INDEX(Stock_Master!J$2:J$5000,{idx}))').border = THIN_BORDER

    tot_row = row + 21
    ws.cell(row=tot_row, column=1, value="TOTALS").font = BOLD_FONT
    ws.cell(row=tot_row, column=1).border = THIN_BORDER
    ws.cell(row=tot_row, column=6, value=f'=SUM(F{row+1}:F{tot_row-1})').font = BOLD_FONT
    ws.cell(row=tot_row, column=6).border = THIN_BORDER
    ws.cell(row=tot_row, column=8, value=f'=SUM(H{row+1}:H{tot_row-1})').font = BOLD_FONT
    ws.cell(row=tot_row, column=8).number_format = FMT_NUM
    ws.cell(row=tot_row, column=8).border = THIN_BORDER

    set_col_widths(ws, [(1, 22), (2, 10), (3, 12), (4, 12), (5, 12), (6, 12), (7, 14), (8, 14), (9, 14)])

    # Conditional formatting
    ws.conditional_formatting.add(f"A{row+1}:I{tot_row-1}", FormulaRule(formula=[f'$I{row+1}="Reorder"'], fill=RED_FILL))
    ws.conditional_formatting.add(f"A{row+1}:I{tot_row-1}", FormulaRule(formula=[f'$I{row+1}="Low Stock"'], fill=YELLOW_FILL))
    ws.conditional_formatting.add(f"A{row+1}:I{tot_row-1}", FormulaRule(formula=[f'$I{row+1}="In Stock"'], fill=GREEN_FILL))


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 10: PRINTABLE INVOICE
# ═══════════════════════════════════════════════════════════════════════════

def build_printable_invoice(ws):
    ws.title = "Printable_Invoice"

    # Business Header
    ws.cell(row=1, column=1, value="[YOUR BUSINESS NAME HERE]").font = Font(name="Calibri", bold=True, size=20, color=DARK_BLUE)
    ws.merge_cells("A1:G1")
    ws.cell(row=2, column=1, value="[Address Line 1]").font = Font(name="Calibri", size=10, color=TEXT_MED)
    ws.merge_cells("A2:G2")
    ws.cell(row=3, column=1, value="[Phone] | [Email] | [PAN/VAT]").font = Font(name="Calibri", size=10, color=TEXT_MED)
    ws.merge_cells("A3:G3")

    # Invoice Title
    ws.cell(row=5, column=1, value="").font = NORMAL_FONT
    ws.merge_cells("A5:G5")
    ws.merge_cells("A6:G6")
    border_cell = ws.cell(row=6, column=1, value="TAX INVOICE")
    border_cell.font = Font(name="Calibri", bold=True, size=16, color=BG_WHITE)
    border_cell.fill = HEADER_FILL
    border_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 36
    for c in range(1, 8):
        ws.cell(row=6, column=c).fill = HEADER_FILL
        ws.cell(row=6, column=c).border = THIN_BORDER

    # Invoice Details
    ws.cell(row=8, column=1, value="Invoice No:").font = BOLD_FONT
    ws.cell(row=8, column=2, value='IFERROR(INDEX(Sales_Entry!B$2:B$5000,MATCH(1,(Sales_Entry!A$2:A$5000=TODAY()),0)),"INV-____")').font = NORMAL_FONT
    ws.cell(row=9, column=1, value="Date:").font = BOLD_FONT
    ws.cell(row=9, column=2, value=f'=TODAY()').number_format = "DD-MMM-YYYY"

    # Customer Details
    ws.cell(row=8, column=4, value="Customer:").font = BOLD_FONT
    ws.merge_cells("D8:G8")
    ws.cell(row=9, column=4, value="[Customer Name]").font = NORMAL_FONT
    ws.cell(row=10, column=4, value="[Address]").font = NORMAL_FONT
    ws.cell(row=11, column=4, value="[Phone]").font = NORMAL_FONT

    # Items Table
    row = 13
    inv_headers = ["#", "Product Name", "Quantity", "Unit", "Rate", "Amount", "Discount", "Net Amount"]
    apply_header_style(ws, inv_headers, row=row)
    for c in range(1, 9):
        ws.cell(row=row, column=c).fill = HEADER_FILL
        ws.cell(row=row, column=c).font = HEADER_FONT

    for i in range(5):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=i+1).border = THIN_BORDER
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=3).border = THIN_BORDER
        ws.cell(row=r, column=4).border = THIN_BORDER
        ws.cell(row=r, column=5).border = THIN_BORDER
        ws.cell(row=r, column=5).number_format = FMT_NUM
        ws.cell(row=r, column=6).border = THIN_BORDER
        ws.cell(row=r, column=6).number_format = FMT_NUM
        ws.cell(row=r, column=7).border = THIN_BORDER
        ws.cell(row=r, column=7).number_format = FMT_NUM
        ws.cell(row=r, column=8).border = THIN_BORDER
        ws.cell(row=r, column=8).number_format = FMT_NUM
        ws.cell(row=r, column=8).font = BOLD_FONT

    # Totals
    total_row = row + 6
    ws.cell(row=total_row, column=7, value="Subtotal:").font = BOLD_FONT
    ws.cell(row=total_row, column=7).border = THIN_BORDER
    ws.cell(row=total_row, column=8, value=f'=SUM(F{row+1}:F{total_row-1})').font = BOLD_FONT
    ws.cell(row=total_row, column=8).number_format = FMT_NUM
    ws.cell(row=total_row, column=8).border = THIN_BORDER

    ws.cell(row=total_row+1, column=7, value="Discount:").font = BOLD_FONT
    ws.cell(row=total_row+1, column=7).border = THIN_BORDER
    ws.cell(row=total_row+1, column=8, value=f'=SUM(G{row+1}:G{total_row-1})').font = BOLD_FONT
    ws.cell(row=total_row+1, column=8).number_format = FMT_NUM
    ws.cell(row=total_row+1, column=8).border = THIN_BORDER

    ws.cell(row=total_row+2, column=7, value="Grand Total:").font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
    ws.cell(row=total_row+2, column=7).border = THIN_BORDER
    ws.cell(row=total_row+2, column=7).fill = LIGHT_BLUE_FILL
    ws.cell(row=total_row+2, column=8, value=f'=SUM(H{row+1}:H{total_row-1})').font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
    ws.cell(row=total_row+2, column=8).number_format = FMT_NUM
    ws.cell(row=total_row+2, column=8).border = THIN_BORDER
    ws.cell(row=total_row+2, column=8).fill = LIGHT_BLUE_FILL
    for c in range(1, 7):
        ws.cell(row=total_row+2, column=c).fill = LIGHT_BLUE_FILL
        ws.cell(row=total_row+2, column=c).border = THIN_BORDER

    # Amount in Words placeholder
    ws.cell(row=total_row+4, column=1, value="Amount in Words:").font = BOLD_FONT
    ws.cell(row=total_row+4, column=2, value="[Rupees ____________________________________]").font = NORMAL_FONT
    ws.merge_cells(f"B{total_row+4}:G{total_row+4}")

    # Signature and Terms
    ws.cell(row=total_row+6, column=1, value="Terms & Conditions:").font = BOLD_FONT
    ws.cell(row=total_row+7, column=1, value="1. Goods once sold will not be taken back.").font = SMALL_FONT
    ws.cell(row=total_row+8, column=1, value="2. Payment due within 15 days from invoice date.").font = SMALL_FONT
    ws.cell(row=total_row+9, column=1, value="3. Interest @ 18% p.a. charged on overdue payments.").font = SMALL_FONT

    ws.cell(row=total_row+6, column=6, value="Authorized Signature").font = BOLD_FONT
    ws.merge_cells(f"F{total_row+6}:H{total_row+6}")
    ws.cell(row=total_row+8, column=6, value="(___________________)").font = NORMAL_FONT

    # Column widths
    set_col_widths(ws, [(1, 6), (2, 28), (3, 10), (4, 10), (5, 12), (6, 14), (7, 14), (8, 16)])

    # Print settings
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 11: PRINTABLE LEDGER STATEMENT


# ═══════════════════════════════════════════════════════════════════════════

def build_printable_ledger(ws):
    ws.title = "Printable_Ledger"

    # Business Header
    ws.cell(row=1, column=1, value="[YOUR BUSINESS NAME HERE]").font = Font(name="Calibri", bold=True, size=20, color=DARK_BLUE)
    ws.merge_cells("A1:H1")
    ws.cell(row=2, column=1, value="[Address] | [Phone] | [PAN/VAT]").font = Font(name="Calibri", size=10, color=TEXT_MED)
    ws.merge_cells("A2:H2")

    # Statement Title
    ws.merge_cells("A4:H4")
    ws.cell(row=4, column=1, value="PARTY LEDGER STATEMENT").font = Font(name="Calibri", bold=True, size=16, color=DARK_BLUE)
    ws.cell(row=4, column=1).alignment = Alignment(horizontal="center")

    # Party Info
    ws.cell(row=6, column=1, value="Party Name:").font = BOLD_FONT
    ws.cell(row=6, column=2, value="[Enter Party Name]").font = NORMAL_FONT
    ws.cell(row=7, column=1, value="Period:").font = BOLD_FONT
    ws.cell(row=7, column=2, value=f'=TEXT(DATE(YEAR(TODAY()),MONTH(TODAY()),1),"DD-MMM-YYYY")&" To "&TEXT(TODAY(),"DD-MMM-YYYY")').font = NORMAL_FONT

    # Opening Balance
    ws.cell(row=9, column=1, value="Opening Balance:").font = BOLD_FONT
    ws.merge_cells("A9:B9")
    ws.cell(row=9, column=3, value=0).number_format = FMT_NUM
    ws.cell(row=9, column=3).font = BOLD_FONT

    # Ledger Table
    row = 11
    ledger_headers = ["Date", "Transaction Type", "Reference", "Description", "Debit", "Credit", "Balance"]
    apply_header_style(ws, ledger_headers, row=row)
    for c in range(1, 8):
        ws.cell(row=row, column=c).fill = HEADER_FILL
        ws.cell(row=row, column=c).font = HEADER_FONT

    for i in range(15):
        r = row + 1 + i
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).number_format = FMT_DATE_SHORT
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=3).border = THIN_BORDER
        ws.cell(row=r, column=4).border = THIN_BORDER
        ws.cell(row=r, column=5).border = THIN_BORDER
        ws.cell(row=r, column=5).number_format = FMT_NUM
        ws.cell(row=r, column=6).border = THIN_BORDER
        ws.cell(row=r, column=6).number_format = FMT_NUM
        ws.cell(row=r, column=7).border = THIN_BORDER
        ws.cell(row=r, column=7).number_format = FMT_NUM
        ws.cell(row=r, column=7).font = BOLD_FONT

    # Totals
    tot_row = row + 16
    ws.cell(row=tot_row, column=4, value="Total").font = BOLD_FONT
    ws.cell(row=tot_row, column=4).border = THIN_BORDER
    ws.cell(row=tot_row, column=5, value=f'=SUM(E{row+1}:E{tot_row-1})').font = BOLD_FONT
    ws.cell(row=tot_row, column=5).number_format = FMT_NUM
    ws.cell(row=tot_row, column=5).border = THIN_BORDER
    ws.cell(row=tot_row, column=6, value=f'=SUM(F{row+1}:F{tot_row-1})').font = BOLD_FONT
    ws.cell(row=tot_row, column=6).number_format = FMT_NUM
    ws.cell(row=tot_row, column=6).border = THIN_BORDER
    ws.cell(row=tot_row, column=7, value=f'=SUM(G{row+1}:G{tot_row-1})').font = BOLD_FONT
    ws.cell(row=tot_row, column=7).number_format = FMT_NUM
    ws.cell(row=tot_row, column=7).border = THIN_BORDER
    for c in range(1, 8):
        ws.cell(row=tot_row, column=c).fill = LIGHT_BLUE_FILL
        ws.cell(row=tot_row, column=c).border = THIN_BORDER

    # Closing Balance
    ws.cell(row=tot_row+2, column=1, value="Closing Balance:").font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
    ws.cell(row=tot_row+2, column=1).border = THIN_BORDER
    ws.cell(row=tot_row+2, column=7, value=f'=C9+E{tot_row}-F{tot_row}').font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
    ws.cell(row=tot_row+2, column=7).number_format = FMT_NUM
    ws.cell(row=tot_row+2, column=7).border = THIN_BORDER
    for c in range(2, 7):
        ws.cell(row=tot_row+2, column=c).border = THIN_BORDER

    # Signature
    ws.cell(row=tot_row+5, column=6, value="Prepared By:").font = BOLD_FONT
    ws.merge_cells(f"F{tot_row+5}:H{tot_row+5}")

    set_col_widths(ws, [(1, 14), (2, 18), (3, 14), (4, 30), (5, 14), (6, 14), (7, 14), (8, 12)])

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 12: SETTINGS / MASTER DATA


# ═══════════════════════════════════════════════════════════════════════════

def build_settings(ws):
    ws.title = "Settings"

    add_title(ws, BUSINESS_NAME, "⚙️ Settings & Master Data Configuration", row=1)

    # Business Information
    row = 4
    ws.cell(row=row, column=1, value="🏢 BUSINESS INFORMATION").font = BOLD_12_FONT
    row += 1
    settings_data = [
        ("Business Name:", "[Your Dairy Plant Name]"),
        ("Address Line 1:", "[Address]"),
        ("Address Line 2:", "[City, District]"),
        ("Phone:", "[Phone Number]"),
        ("Email:", "[Email Address]"),
        ("PAN / VAT No:", "[PAN Number]"),
        ("Registration No:", "[Registration No]"),
        ("Currency:", "NPR (Rs.)"),
    ]
    for i, (label, value) in enumerate(settings_data):
        r = row + i
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=value).font = NORMAL_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.merge_cells(f"B{r}:D{r}")

    # Units Master
    row = row + len(settings_data) + 2
    ws.cell(row=row, column=1, value="📏 UNITS OF MEASUREMENT").font = BOLD_12_FONT
    row += 1
    units = ["Liter", "Kg", "Gram", "Piece", "Packet", "Box", "Carton", "Tin"]
    apply_header_style(ws, ["#", "Unit Name"], row=row)
    for i, u in enumerate(units):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=i+1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=u).font = NORMAL_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER

    # Payment Modes
    row = row + 1 + len(units) + 1
    ws.cell(row=row, column=1, value="💳 PAYMENT MODES").font = BOLD_12_FONT
    row += 1
    modes = ["Cash", "Credit", "Bank Transfer", "UPI", "Cheque", "Card", "Online"]
    apply_header_style(ws, ["#", "Payment Mode"], row=row)
    for i, m in enumerate(modes):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=i+1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=m).font = NORMAL_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER

    # Transaction Types
    row = row + 1 + len(modes) + 1
    ws.cell(row=row, column=1, value="📋 TRANSACTION TYPES").font = BOLD_12_FONT
    row += 1
    types = ["Sale", "Purchase", "Receipt", "Payment", "Opening", "Adjustment", "Credit Note", "Debit Note"]
    apply_header_style(ws, ["#", "Transaction Type"], row=row)
    for i, t in enumerate(types):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=i+1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=t).font = NORMAL_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER

    # Product Categories
    row = row + 1 + len(types) + 1
    ws.cell(row=row, column=1, value="📦 PRODUCT CATEGORIES").font = BOLD_12_FONT
    row += 1
    cats = ["Milk", "Curd", "Ghee", "Paneer", "Butter", "Buttermilk", "Khoya", "Other"]
    apply_header_style(ws, ["#", "Category"], row=row)
    for i, c in enumerate(cats):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=i+1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=c).font = NORMAL_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER

    # Status Options
    row = row + 1 + len(cats) + 1
    ws.cell(row=row, column=1, value="✅ STATUS OPTIONS").font = BOLD_12_FONT
    row += 1
    sts = ["Paid", "Unpaid", "Partial", "Pending", "Processed", "Cancelled"]
    apply_header_style(ws, ["#", "Status"], row=row)
    for i, s in enumerate(sts):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=i+1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=s).font = NORMAL_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER

    # Milk Pricing Constants
    row = row + 1 + len(sts) + 2
    ws.cell(row=row, column=1, value="🧮 MILK PRICING CONSTANTS").font = BOLD_12_FONT
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    pricing_info = [
        ("FAT Rate Multiplier:", 7.15),
        ("SNF Rate Multiplier:", 4.55),
        ("Rate Type Options:", "FORMULA, FIXED"),
        ("Pricing Formula (FORMULA):", "Rate = (FAT% × 7.15) + (SNF% × 4.55) + Extra Per Unit"),
        ("Pricing Formula (FIXED):", "Rate = Fixed Rate Per Unit"),
        ("Amount Formula:", "Amount = Quantity × Rate Per Unit"),
        ("Net Amount:", "Amount + Transport Charges"),
    ]
    for i, (label, value) in enumerate(pricing_info):
        r = row + i
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=str(value)).font = NORMAL_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.merge_cells(f"B{r}:H{r}")

    # Quick Reference
    row = row + 1 + len(sts) + 2
    ws.cell(row=row, column=1, value="📖 QUICK REFERENCE").font = BOLD_12_FONT
    row += 1
    refs = [
        "Sales Entry: Record all sales invoices in the Sales_Entry sheet.",
        "Purchase Entry: Record all purchase bills in the Purchase_Entry sheet.",
        "Stock Master: Products auto-update from sales and purchases.",
        "Party Master: Customer/supplier balances auto-calculate.",
        "Party Ledger: Running balance ledger for each party.",
        "Dashboard: All KPIs and summaries auto-calculate.",
        "Reports: Filterable sales/purchase/stock reports.",
        "Printable Invoice: Professional invoice for printing.",
        "Printable Ledger: Statement for party-wise balance.",
        "",
        "💡 TIP: Change the business name in the Dashboard, Printable_Invoice, and Printable_Ledger sheets.",
        "💡 TIP: Add new items to the lists above to expand dropdown options in data sheets.",
    ]
    for i, ref in enumerate(refs):
        ws.cell(row=row + i, column=1, value=ref).font = NORMAL_FONT
        ws.merge_cells(f"A{row+i}:H{row+i}")

    set_col_widths(ws, [(1, 28), (2, 30), (3, 18), (4, 18), (5, 18), (6, 18), (7, 18), (8, 18)])

    # Protect the settings sheet
    ws.protection.sheet = False  # Keep it editable for now

    ws.page_setup.orientation = "portrait"


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 13: DAYBOOK (Traditional Debit/Credit Accounting)
# ═══════════════════════════════════════════════════════════════════════════

def build_daybook(ws):
    ws.title = "Daybook"

    add_title(ws, BUSINESS_NAME, "📒 Daybook — Daily Debit/Credit Register", row=1)

    ws.cell(row=4, column=1, value="Instructions:").font = BOLD_12_FONT
    ws.merge_cells("A4:H4")
    ws.cell(row=5, column=1, value="Record ALL daily transactions here. Debit = money outflow (expenses, purchases), Credit = money inflow (sales, receipts). Balance updates automatically.").font = SMALL_FONT
    ws.merge_cells("A5:H5")
    ws.cell(row=6, column=1, value="💡 Sales & Purchase entries can be manually posted here from their respective sheets for a complete audit trail.").font = SMALL_FONT
    ws.merge_cells("A6:H6")

    headers = [
        "Date", "Particulars", "Voucher Type", "Reference No",
        "Debit (Dr)", "Credit (Cr)", "Balance", "Narration"
    ]

    # Sample daybook entries
    data = [
        [datetime.date(2026, 3, 1), "Opening Balance", "Opening", "", "", 50000, 50000, "Opening cash & bank balance"],
        [TODAY - datetime.timedelta(days=10), "Sharma Dairy Shop", "Sale", "INV-1001", "", 5800, 55800, "Sale of milk & products"],
        [TODAY - datetime.timedelta(days=9), "Gurung Dairy Farm", "Purchase", "BILL-5001", 15300, "", 40500, "Milk purchase with transport"],
        [TODAY - datetime.timedelta(days=8), "Cash Sales", "Sale", "CS-001", "", 4200, 44700, "Counter sales - daily collection"],
        [TODAY - datetime.timedelta(days=7), "Salary", "Payment", "PAY-001", 12000, "", 32700, "Monthly staff salary"],
        [TODAY - datetime.timedelta(days=6), "Bhatbhateni Supermarket", "Sale", "INV-1002", "", 8500, 41200, "Bulk ghee & paneer supply"],
        [TODAY - datetime.timedelta(days=5), "Transport", "Payment", "PAY-002", 2500, "", 38700, "Milk transport charges"],
        [TODAY - datetime.timedelta(days=4), "Sharma Dairy Shop", "Receipt", "PAY-101", "", 10000, 48700, "Payment received - bank transfer"],
        [TODAY - datetime.timedelta(days=3), "Electricity", "Payment", "PAY-003", 3500, "", 45200, "Monthly electricity bill"],
        [TODAY - datetime.timedelta(days=2), "Fresh Valley Suppliers", "Purchase", "BILL-5002", 12500, "", 32700, "Milk purchase - Formula rate"],
        [TODAY - datetime.timedelta(days=1), "Hotel Annapurna", "Sale", "INV-1003", "", 3200, 35900, "Ghee & butter supply"],
    ]

    add_excel_table(ws, headers, data, "DaybookTable", start_row=8,
                    col_widths=[(1, 14), (2, 26), (3, 16), (4, 14), (5, 14), (6, 14), (7, 14), (8, 36)])

    # Balance formula with carry-forward (col G = column 7)
    # First data row (r=9) is the opening balance — keep its static value
    # Balance formulas start from second data row (r=10)
    for r in range(10, 9 + len(data)):
        # Balance = Previous Balance + Credit - Debit
        ws.cell(row=r, column=7).value = f'=G{r-1}+IF(F{r}="",0,F{r})-IF(E{r}="",0,E{r})'
        ws.cell(row=r, column=7).number_format = FMT_NUM
        ws.cell(row=r, column=7).font = BOLD_FONT

    # Totals row
    tot_row = 9 + len(data)
    ws.cell(row=tot_row, column=3, value="TOTALS").font = BOLD_FONT
    ws.cell(row=tot_row, column=3).border = THIN_BORDER
    ws.cell(row=tot_row, column=5, value=f'=SUM(E9:E{tot_row-1})').font = BOLD_FONT
    ws.cell(row=tot_row, column=5).number_format = FMT_NUM
    ws.cell(row=tot_row, column=5).border = THIN_BORDER
    ws.cell(row=tot_row, column=5).fill = LIGHT_BLUE_FILL
    ws.cell(row=tot_row, column=6, value=f'=SUM(F9:F{tot_row-1})').font = BOLD_FONT
    ws.cell(row=tot_row, column=6).number_format = FMT_NUM
    ws.cell(row=tot_row, column=6).border = THIN_BORDER
    ws.cell(row=tot_row, column=6).fill = LIGHT_BLUE_FILL
    for c in range(1, 9):
        ws.cell(row=tot_row, column=c).fill = LIGHT_BLUE_FILL
        ws.cell(row=tot_row, column=c).border = THIN_BORDER

    # Data validation for Voucher Type
    add_data_validation(ws, "C", 9, 5000,
                        ["Sale", "Purchase", "Receipt", "Payment", "Opening", "Expense", "Income", "Transfer", "Adjustment"],
                        "Voucher Type", "Select transaction type")

    # Conditional formatting (debit high = red, credit high = green for the row)
    ws.conditional_formatting.add("A9:H5000", FormulaRule(formula=['$E9>0'], fill=RED_FILL))
    ws.conditional_formatting.add("A9:H5000", FormulaRule(formula=['$F9>0'], fill=GREEN_FILL))
    ws.conditional_formatting.add("G9:G5000", CellIsRule(operator="lessThan", formula=["0"], fill=YELLOW_FILL))

    # Freeze panes at data start
    ws.freeze_panes = ws.cell(row=9, column=1).coordinate

    ws.page_setup.orientation = "landscape"


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 14: CASH COLLECTION (from Daybook_Accounting_System)
# ═══════════════════════════════════════════════════════════════════════════

def build_cash_collection(ws):
    ws.title = "Cash_Collection"

    headers = [
        "Date", "Receipt No", "Customer Name", "Against Bill No",
        "Opening Due", "Collected Amount", "Payment Mode", "Closing Due", "Remarks"
    ]

    data = [
        [TODAY - datetime.timedelta(days=10), "R001", "Sharma Dairy Shop", "INV-1001", 5800, 5800, "Cash", 0, "Full payment received"],
        [TODAY - datetime.timedelta(days=8), "R002", "Krishna Store", "INV-1004", 3960, 3960, "Cash", 0, "Full payment"],
        [TODAY - datetime.timedelta(days=6), "R003", "Patan Co-op", "INV-1005", 7200, 5000, "Bank Transfer", 2200, "Partial payment"],
        [TODAY - datetime.timedelta(days=5), "R004", "Bhatbhateni Supermarket", "INV-1002", 8500, 8500, "Cheque", 0, "Full payment via cheque"],
        [TODAY - datetime.timedelta(days=3), "R005", "Hotel Annapurna", "INV-1003", 3200, 3200, "Cash", 0, ""],
        [TODAY - datetime.timedelta(days=1), "R006", "Patan Co-op", "INV-1005", 2200, 2200, "UPI", 0, "Balance cleared"],
    ]

    add_excel_table(ws, headers, data, "CashCollectionTable", start_row=1,
                    col_widths=[(1, 14), (2, 14), (3, 24), (4, 16), (5, 14), (6, 16), (7, 16), (8, 14), (9, 24)])

    # Closing Due formula (col H = column 8)
    for r in range(2, 8):
        ws.cell(row=r, column=8).value = f'=E{r}-F{r}'
        ws.cell(row=r, column=8).number_format = FMT_NUM
        ws.cell(row=r, column=8).font = BOLD_FONT

    # Data validations
    add_data_validation(ws, "G", 2, 5000, ["Cash", "Bank Transfer", "Cheque", "UPI", "Card", "Online"], "Payment Mode", "Select payment mode")

    # Conditional formatting
    ws.conditional_formatting.add("A2:I5000", FormulaRule(formula=['$H2>0'], fill=YELLOW_FILL))
    ws.conditional_formatting.add("A2:I5000", FormulaRule(formula=['$H2=0'], fill=GREEN_FILL))


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 15: DENOMINATION COUNT (from Daybook_Accounting_System)
# ═══════════════════════════════════════════════════════════════════════════

def build_denomination_count(ws):
    ws.title = "Denomination_Count"

    add_title(ws, BUSINESS_NAME, "💰 Daily Cash Denomination Count", row=1)

    headers = [
        "Date", "Count ₹5", "Count ₹10", "Count ₹20", "Count ₹50",
        "Count ₹100", "Count ₹500", "Count ₹1000", "Other Amount", "Grand Total"
    ]

    data = []
    for i in range(10):
        d = TODAY - datetime.timedelta(days=i)
        c5 = i * 5 + 10
        c10 = i * 3 + 15
        c20 = i * 2 + 8
        c50 = i * 1 + 5
        c100 = i * 2 + 12
        c500 = i * 1 + 3
        c1000 = i * 1 + 2
        other = round(50 + i * 15, 0)
        data.append([d, c5, c10, c20, c50, c100, c500, c1000, other, 0])

    add_excel_table(ws, headers, data, "DenomTable", start_row=3,
                    col_widths=[(1, 14), (2, 12), (3, 12), (4, 12), (5, 12), (6, 12), (7, 12), (8, 12), (9, 14), (10, 14)])

    # Grand Total formula (col J = column 10)
    for r in range(4, 4 + len(data)):
        ws.cell(row=r, column=10).value = f'=B{r}*5 + C{r}*10 + D{r}*20 + E{r}*50 + F{r}*100 + G{r}*500 + H{r}*1000 + I{r}'
        ws.cell(row=r, column=10).number_format = FMT_NUM_INT
        ws.cell(row=r, column=10).font = BOLD_FONT

    # Totals row
    tot_row = 4 + len(data)
    ws.cell(row=tot_row, column=1, value="TOTALS").font = BOLD_FONT
    ws.cell(row=tot_row, column=1).border = THIN_BORDER
    for c in range(2, 11):
        ws.cell(row=tot_row, column=c, value=f'=SUM({get_column_letter(c)}4:{get_column_letter(c)}{tot_row-1})').font = BOLD_FONT
        ws.cell(row=tot_row, column=c).number_format = FMT_NUM_INT
        ws.cell(row=tot_row, column=c).border = THIN_BORDER
        ws.cell(row=tot_row, column=c).fill = LIGHT_BLUE_FILL
    for c in range(1, 11):
        ws.cell(row=tot_row, column=c).fill = LIGHT_BLUE_FILL
        ws.cell(row=tot_row, column=c).border = THIN_BORDER


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 16: README / INFO
# ═══════════════════════════════════════════════════════════════════════════

def build_readme(ws):
    ws.title = "README"

    ws.merge_cells("A1:H1")
    ws.cell(row=1, column=1, value=f"🐄 {BUSINESS_NAME} — User Guide & Reference").font = TITLE_FONT

    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1, value=f"Generated on {TODAY.strftime('%d-%b-%Y')} | Version {VERSION} | Merged Workbook (Dairy Plant + Daybook)").font = SUBTITLE_FONT

    row = 4
    sections = [
        ("📋 SHEET PURPOSES", [
            ("Dashboard", "Main overview: KPIs, monthly trends, party balances, low stock alerts, recent transactions. Navigation links to all sheets."),
            ("Daybook", "Daily debit/credit register. Record all transactions here. Debit = outflow (expenses), Credit = inflow (income). Running balance auto-calculates."),
            ("Sales_Entry", "Record all sales invoices. Products, quantities, rates, discounts, and payment status. Auto-updates stock via SUMIF formulas."),
            ("Purchase_Entry", "Record milk/raw material purchases. Uses component-based pricing: (FAT% × 7.15) + (SNF% × 4.55) + Extra. Or use fixed rate per liter."),
            ("Cash_Collection", "Track customer payments against invoices. Shows opening due, collected amount, and closing balance."),
            ("Stock_Master", "Product inventory management. Opening stock + Purchases - Sales = Current stock. Auto-calculates stock value and reorder status."),
            ("Party_Master", "Customer/supplier directory with contact info, opening balances, and auto-calculated running balances."),
            ("Party_Ledger", "Party-wise transaction ledger: all sales, purchases, receipts, and payments for each party with running balance."),
            ("Denomination_Count", "Daily cash counting by denomination (₹5 to ₹1000). Use for end-of-day cash reconciliation."),
            ("Sales_Report", "Filterable sales report. Set date range and party for summary."),
            ("Purchase_Report", "Filterable purchase report with same controls."),
            ("Stock_Report", "Stock summary with valuation and reorder alerts."),
            ("Printable_Invoice", "Professional tax invoice for printing/saving as PDF."),
            ("Printable_Ledger", "Party ledger statement for printing/saving as PDF."),
            ("Settings", "Business info, master data lists, and configuration constants."),
            ("README", "This guide."),
        ]),
        ("🧮 MILK PURCHASE PRICING", [
            ("Formula Rate:", "Rate = (FAT% × 7.15) + (SNF% × 4.55) + Extra Per Unit"),
            ("Fixed Rate:", "Rate = Fixed Rate Per Unit (simple per-liter price)"),
            ("Switching:", "Change Rate Type dropdown to FORMULA or FIXED in Purchase_Entry"),
            ("Extra Per Unit:", "Contract-based addition per liter (e.g., +2 for quality premium)"),
            ("Example:", "FAT=3.5, SNF=8.5, Extra=2 → Rate = (3.5×7.15)+(8.5×4.55)+2 = 25.03+38.68+2 = 65.70"),
        ]),
        ("📒 HOW TO USE THE DAYBOOK", [
            ("Step 1:", "Enter Date and Particulars (account/party name)"),
            ("Step 2:", "Select Voucher Type from dropdown (Sale, Purchase, Receipt, Payment, etc.)"),
            ("Step 3:", "Enter Reference No (invoice/bill/payment number)"),
            ("Step 4:", "Enter amount in Debit (expense) or Credit (income) column"),
            ("Step 5:", "Add a Narration/description"),
            ("Balance:", "Auto-calculates as: Previous Balance + Credit - Debit"),
        ]),
        ("🔄 HOW SHEETS INTERACT", [
            ("Sales → Stock:", "Sales_Entry quantities auto-subtract from Stock_Master via SUMIF"),
            ("Purchases → Stock:", "Purchase_Entry quantities auto-add to Stock_Master via SUMIF"),
            ("Sales → Daybook:", "Optionally post sales as Credit entries in Daybook"),
            ("Purchases → Daybook:", "Optionally post purchases as Debit entries in Daybook"),
            ("Cash Collection:", "Receipts reduce outstanding dues from customers"),
            ("Party Ledger:", "Auto-aggregates party balances from transactions"),
            ("Dashboard:", "All KPIs pull from data sheets automatically"),
        ]),
        ("📊 CHECKING TOTALS & BALANCES", [
            ("Monthly Sales:", "Dashboard KPI card shows current month's total sales"),
            ("Monthly Purchases:", "Dashboard KPI card shows current month's total purchases"),
            ("Stock Value:", "Dashboard KPI card shows total inventory value"),
            ("Party Balances:", "Dashboard → Party Balance Summary section"),
            ("Daybook Balance:", "Daybook last row shows running cash/bank balance"),
            ("Low Stock:", "Dashboard → Low Stock Alerts section"),
        ]),
    ]

    for section_title, items in sections:
        ws.cell(row=row, column=1, value=section_title).font = BOLD_12_FONT
        ws.merge_cells(f"A{row}:H{row}")
        ws.row_dimensions[row].height = 28
        row += 1

        for label, desc in items:
            ws.cell(row=row, column=1, value=label).font = BOLD_FONT
            ws.cell(row=row, column=2, value=desc).font = NORMAL_FONT
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"B{row}:H{row}")
            ws.row_dimensions[row].height = 22
            row += 1

        row += 1  # spacing

    set_col_widths(ws, [(1, 24), (2, 18), (3, 18), (4, 18), (5, 18), (6, 18), (7, 18), (8, 18)])
    ws.page_setup.orientation = "portrait"


# ═══════════════════════════════════════════════════════════════════════════
# MAIN GENERATION
# ═══════════════════════════════════════════════════════════════════════════

def generate_workbooks():
    print("🐄 Dairy Plant Accounts — Professional Workbook Generator")
    print("=" * 60)

    # ── VERSION 1: Formula-Only (.xlsx) ────────────────────────────────
    print("\n📗 Generating Formula-Only Version (.xlsx)...")
    wb = Workbook()

    sheet_builders = [
        ("Dashboard", build_dashboard),
        ("Daybook", build_daybook),
        ("Sales_Entry", build_sales_entry),
        ("Purchase_Entry", build_purchase_entry),
        ("Cash_Collection", build_cash_collection),
        ("Stock_Master", build_stock_master),
        ("Party_Master", build_party_master),
        ("Party_Ledger", build_party_ledger),
        ("Denomination_Count", build_denomination_count),
        ("Sales_Report", build_sales_report),
        ("Purchase_Report", build_purchase_report),
        ("Stock_Report", build_stock_report),
        ("Printable_Invoice", build_printable_invoice),
        ("Printable_Ledger", build_printable_ledger),
        ("Settings", build_settings),
        ("README", build_readme),
    ]

    ws0 = wb.active
    ws0.title = "temp"

    for idx, (name, builder) in enumerate(sheet_builders):
        if idx == 0:
            ws = ws0
            ws.title = name
        else:
            ws = wb.create_sheet(title=name)
        print(f"  📄 Building sheet: {name}")
        builder(ws)

    # Set active to Dashboard
    wb.active = 0  # Dashboard is first sheet

    # Tab colors for visual organization
    tab_colors = {
        "Dashboard": "1B3A5C",
        "Daybook": "2ECC71",
        "Sales_Entry": "27AE60",
        "Purchase_Entry": "E74C3C",
        "Cash_Collection": "3498DB",
        "Stock_Master": "F39C12",
        "Party_Master": "8E44AD",
        "Party_Ledger": "2C3E50",
        "Denomination_Count": "1ABC9C",
        "Sales_Report": "1ABC9C",
        "Purchase_Report": "E67E22",
        "Stock_Report": "3498DB",
        "Printable_Invoice": "9B59B6",
        "Printable_Ledger": "34495E",
        "Settings": "7F8C8D",
        "README": "95A5A6",
    }
    for name, color in tab_colors.items():
        try:
            wb[name].sheet_properties.tabColor = color
        except KeyError:
            pass

    wb.save(OUTPUT_XLSX)
    print(f"\n✅ Formula-Only version saved: {OUTPUT_XLSX}")
    print(f"   File size: {os.path.getsize(OUTPUT_XLSX) / 1024:.1f} KB")

    # ── VERSION 2: Create .xlsm placeholder ────────────────────────────
    # openpyxl can save as .xlsm but can't create VBA code.
    # The .bas file is provided separately.
    # We create the .xlsm as a copy of .xlsx for VBA import.
    OUTPUT_XLSM = os.path.join(OUTPUT_DIR, "Dairy_Plant_Accounts_VBA.xlsm")
    wb.save(OUTPUT_XLSM)
    print(f"\n✅ VBA version placeholder saved: {OUTPUT_XLSM}")
    print(f"   ⚠️  Import the VBA macros file (Dairy_Plant_VBA_Macros.bas) into this workbook.")
    print(f"   ⚠️  Instructions: Open Excel > Alt+F11 > File > Import File > Select .bas file")

    print(f"\n📊 Workbook Summary:")
    print(f"   Sheets: {', '.join(name for name, _ in sheet_builders)}")
    print(f"   Total Sheets: {len(sheet_builders)}")
    print(f"   Compatibility: Windows 7 + Office 2007+, macOS, LibreOffice")
    print(f"   Key Formulas: SUMIFS, INDEX-MATCH, VLOOKUP, IF, SUMPRODUCT")
    print(f"   Features: Data Validation, Conditional Formatting, Freeze Panes, Auto-filters")
    print(f"\n🎉 DONE!")


if __name__ == "__main__":
    generate_workbooks()
